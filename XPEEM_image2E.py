# -*- coding: utf-8 -*-
"""
Module to handle XPEEM 2E-images

Created on Thu Jul 10 14:25:50 2025

@author: lelotte_b
"""

# Basics
import time
import numpy as np
import math

# Display in python
import matplotlib.pyplot as plt
import seaborn as sns 
sns.set_context('paper', font_scale=2.2) # to change the font settings in the graphs
import matplotlib_inline
matplotlib_inline.backend_inline.set_matplotlib_formats('retina') # High resolution plots

# Handle excel inputs and outputs
import pandas as pd
from openpyxl import load_workbook # to delete a sheet easily

# Docstrings, test and warnings
from typing import List, Optional, Tuple, Union
import pytest
import tempfile
np.seterr(all='warn') # to be able to handle warnings
import warnings

# Folder/system management
import psutil
import os
import subprocess


def is_program_running(program_name):
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] == program_name:
            return True
    return False

# Handle image inputs and outputs
from PIL import Image # Simple, Python-native API for image I/O and basic transforms 
import cv2 # fast C++ back-end, huge computer-vision toolset 
import imageio as io # Unified I/O for images, volumes, GIFs, videos 

# Statistics
import scipy.stats as stats
from scipy.stats import linregress, rankdata
import statistics

# Image processing
from skimage.metrics import structural_similarity as ssim
from skimage.metrics import mean_squared_error
from skimage import img_as_ubyte
import statsmodels.api as sm
from scipy.ndimage import gaussian_filter

# My modulesES
import XPEEM
import XPEEM_utils as utils
import OriginPlot as oplt


def PEEM_image(imarray, details=False, sampling=200):
    """
    Process a PEEM image and optionally plot details.

    Parameters
    ----------
    imarray : ndarray
        The input image.
    details : bool, optional
        Whether to plot details. Default is False.
    sampling : int, optional
        The sampling rate. Default is 200.

    Returns
    -------
    tuple
        A tuple containing the processed image, the DIV image, and the min and max y values.
    """
    # Get the size of the image
    n, m = imarray.shape

    # Calculate the x positions in micrometers
    position = np.arange(n)
    x = position * 25 / math.sqrt(2) / n

    # Initialize arrays for the processed images
    imarray_k = np.zeros((n, m))
    imarray_dy = np.zeros((n, m))

    # Process the image values to be within an acceptable range
    imarray, minIm, maxIm, median = acceptable_values(imarray)

    # Process each row of the image
    for i in range(n):
        y = imarray[i, :]
        y = dead_pixel(y, [minIm / 5, maxIm * 5], median)
        imarray_k[i, :] = y

    # Calculate the max and min y values for the image
    max_y = maxIm / 1.5
    min_y = minIm

    # Calculate the DIV image
    DIV = imarray_k / imarray

    # If details are requested, plot the images
    if details:
        utils.plot_images(x, imarray, imarray_dy, DIV, max_y, min_y)

    return (imarray_k, DIV, max_y, min_y)

def test_PEEM_image():
    # Create a synthetic image
    imarray = np.random.randint(0, 256, (100, 100), dtype=np.uint8)

    # Run the function on the synthetic image
    result = PEEM_image(imarray)

    # Check that the output is a tuple of four elements
    assert isinstance(result, tuple)
    assert len(result) == 4

    # Check that the first two elements of the output are numpy arrays of the same shape as the input
    assert isinstance(result[0], np.ndarray)
    assert result[0].shape == imarray.shape
    assert isinstance(result[1], np.ndarray)
    assert result[1].shape == imarray.shape

    # Check that the last two elements of the output are floats
    assert isinstance(result[2], float)
    assert isinstance(result[3], float)

def find_2E(row_args):
    """
    Enumerates through all the "Edge" folders, opens the 2_energy folder for each sample to compare,
    and calls the function PEEM_2i_comp_save.

    Note: the XPEEM dataset should be stored as //path_Project/name_Sample/name_Edge/...
    In the name_Edge folder, one should find the processed E-stack folder and the 2_energies folder.

    Args:
        row_args (pd.Series): A row from a DataFrame containing the 'name_Sample', 'name_ChemAB', 'name_ChemBA', 'path_imA', and 'path_imB' fields.

    Returns:
        paths_2E_AB (list): the path of the two images to process.
    """
    paths_2E_AB = []
    
    pathIms = [f'{row_args["path_imA"]}', f'{row_args["path_imB"]}']
    name_Sample = row_args['name_Sample']
    
    path_Dataset=utils.path_join(os.getcwd(),name_Sample)    

    # Find the 2_energies folder corresponding to this ID.
    for pathIm in pathIms:
        for path_Dataset_Edge in os.scandir(path_Dataset):
            if path_Dataset_Edge.is_dir() and name_Sample.lower() in os.path.basename(path_Dataset_Edge.path).lower():
                # Open the 2_energy folder for each sample to compare
                # List all folders in "2_energies"
                path_2_energies = utils.path_join(path_Dataset_Edge.path, "2_energies")
                folders_2_energies = os.listdir(path_2_energies)
                # Find the folder whose name contains the "sample"
                for folder_2E_i in folders_2_energies:
                    if pathIm  in folder_2E_i and ('_filtered' in folder_2E_i): # and not ('_undistrdd0' in subfolder) :
                        path_2E_i=utils.path_join(path_2_energies, folder_2E_i)
                        if os.path.isdir(path_2E_i):
                            paths_2E_AB.append(path_2E_i)
                            
    assert len(paths_2E_AB)==2, f'The images of the sample {name_Sample} {pathIms[0]} and {pathIms[1]} were been found or there are too many possibilities.'

    return paths_2E_AB

def create_outputFd(path_2E_A: str, path_2E_B: str, row_args: pd.Series, labels: list[str] = ['A', 'B']):
    """
    Compare two sets of images.

    Parameters
    ----------
    folder_2E_A (str):              Path to the first set of images.
    folder_2E_B (str):              Path to the second set of images.
    row_args (pd.Series):           A row from a DataFrame containing the 'Sample', 'name_ChemAB', 'name_ChemBA', 'path_imA', and 'path_imB' fields.
    labels (list of str, optional): Labels for the two sets of images. Default is ['A', 'B'].

    Returns
    -------
    None
    """
    print(row_args['name_ChemAB'])
    
    start_time = time.time()
    
    # Extract the necessary values from the row
    name_Sample = row_args['name_Sample']
    name_Short = oplt.oricompa(name_Sample)
    name_Short = name_Short if len(name_Short)<5 else name_Short[:4]
    name_Chems = [row_args['name_ChemAB'], row_args['name_ChemBA']]
    name_ROIs = row_args['name_ROIs']
    path_ims = [row_args['path_imA'], row_args['path_imB']]
    arg_SourceIm=row_args['arg_SourceIm']
    folder_Origin= '2_i2' if arg_SourceIm =='i2' else '2_ES'
    arg_Analysis = row_args["arg_Analysis"]
    
    # Initialise base directory
    if arg_Analysis == 'Co-localisation':
        folder_Output = arg_SourceIm+'_2E_coloc'
    elif arg_Analysis == 'Ratio' or arg_Analysis == 'Quantification' :
        folder_Output = arg_SourceIm+'_2E_ratio'
    else: 
        raise ValueError('Analysis keyword was not implemented. Analysis should be Co-localisation, Ratio or Quantification')
    
    
    def load_ES_at_E(path_dataset, Etarget, Eref):        
        # Load edge
        edge=utils.find_edge(os.path.basename(path_dataset),exp='SIM')
        
        # Load energies
        E = utils.load_E_I0(path_dataset,processed=True)[0]
        indexE=utils.find_closest_index(E,Etarget)

        if Eref : 
            indexEref=utils.find_closest_index(E,Eref)
        
        # Get the stack folder name and load the image
        folders_dataset = os.listdir(path_dataset)
        folder_dataset_ES=None
        for folder_dataset in folders_dataset:
            if "_processed_test" in folder_dataset :
                folder_dataset_ES=folder_dataset
                path_dataset_ES=utils.path_join(path_dataset,folder_dataset_ES)
                break
        if not folder_dataset_ES :
            raise ValueError
            
        # Import image at energy
        halfwindow=2
        listImsStack = [im for im in os.listdir(path_dataset_ES) if im.endswith('.tif')]
        imNames = listImsStack[indexE-halfwindow:indexE+halfwindow+1]    
        im=np.array([utils.open_image(utils.path_join(path_dataset_ES,imName,dt='f'),format_file='') for imName in imNames])
        if not pd.isnull(Eref) : 
            imRefNames = listImsStack[indexEref-halfwindow:indexEref+halfwindow+1]     
            imref = np.array([utils.open_image(utils.path_join(path_dataset_ES,imRefName,dt='f'),format_file='') for imRefName in imRefNames])
            im /= imref
            
            unrealistic_vals= (im>50) | (im < 0)
            im[unrealistic_vals]=0
        
        return im, edge

    # Load and process images
    if arg_SourceIm == 'i2':
        path_Dataset_A=os.path.abspath(os.path.join(path_2E_A, '..', '..'))
        path_Dataset_B=os.path.abspath(os.path.join(path_2E_B, '..', '..'))
        
        images_A = np.transpose(utils.open_sequence(path_2E_A)[0],(2,0,1))
        images_B = np.transpose(utils.open_sequence(path_2E_B)[0],(2,0,1))
    elif arg_SourceIm == 'ES':
        path_Dataset_A=os.path.abspath(os.path.join(path_2E_A, '..', '..'))
        path_Dataset_B=os.path.abspath(os.path.join(path_2E_B, '..', '..'))
        
        images_A = load_ES_at_E(path_Dataset_A,row_args['E_A'],row_args.get('Eref_A',None))[0]
        images_B = load_ES_at_E(path_Dataset_B,row_args['E_B'],row_args.get('Eref_B',None))[0]
    
    else: 
        raise ValueError('Source image should be 2 Energy Images (i2) or Energy Stack (ES).')
    
    inputFd_path = utils.path_join(os.getcwd(),'_Input')
    outputFd_path = utils.path_join(os.getcwd(),'_Output')
    outputFd2E_path = utils.path_join(outputFd_path, name_Sample, f'{arg_SourceIm}/'+folder_Output)
    
    # Load obtained broad masks from Excel
    list_ROIs, _, mask_ROIs = XPEEM.load_masks(name_Sample, kind='Material')
    
    # Path to the Excel file
    path_comparisons_list = utils.path_join(inputFd_path,'2_args_Maps.xlsx',dt='f')
    
    # If the Excel file exists, read it into a DataFrame. Otherwise, create an empty DataFrame.
    identifiers=['name_Sample', 'name_ROI', 'name_ChemAB']
    try:
        # If the Excel file exists, try to read the "mainfolder" sheet into a DataFrame
        if os.path.exists(path_comparisons_list):
            df = pd.read_excel(path_comparisons_list, sheet_name=folder_Output)
        else:
            raise FileNotFoundError
    except (FileNotFoundError, ValueError):
        # If the "mainfolder" sheet doesn't exist, or if the Excel file doesn't exist, create an empty DataFrame
        df = pd.DataFrame(columns=identifiers)
    
    def create_directories(path_Output, label, subfolders, maskfolders):
        comp_dir = utils.path_join(path_Output, label)
        os.makedirs(comp_dir, exist_ok=True)
    
        # Create subfolders if they do not exist
        for subfolder in subfolders:
            os.makedirs(utils.path_join(comp_dir, subfolder), exist_ok=True)
    
        for maskfolder in maskfolders:
            mskfolder_dir = utils.path_join(comp_dir, 'mask', maskfolder)
            os.makedirs(mskfolder_dir, exist_ok=True)
    
    
    # Create a "material" folder if it does not exist
    for name_ROI, mask_ROI in zip(list_ROIs,mask_ROIs) :
        
        bool_isIncluded=(name_ROI in name_ROIs or name_ROIs == 'Any' and not name_ROI == 'All')
        
        if bool_isIncluded and not arg_Analysis=='Quantification' :
            path_Output_ROI = utils.path_join(outputFd2E_path, name_ROI)
            os.makedirs(path_Output_ROI, exist_ok=True)
            
            # Create a subfolder with the names of the two files being compared
            if row_args["Process"]: name_comparison = f"{name_Chems[0]}_{path_ims[0]}_{path_ims[1]}"
            if row_args["Process_rev"]: name_comparison_rev = f"{name_Chems[1]}_{path_ims[1]}_{path_ims[0]}"
            
            if arg_Analysis == 'Ratio': folders_output_ROI = ["raw", "color", "grayscale_cc"]
            elif arg_Analysis == 'Co-localisation': folders_output_ROI = ["Overlay"]
                
            folders_output_ROI_mask = [f'mask_{trsh}pt' for trsh in [1, 2, 5, 10]]
            
            if row_args["Process"]: create_directories(path_Output_ROI, name_comparison, folders_output_ROI, folders_output_ROI_mask)     
            if row_args["Process_rev"]: create_directories(path_Output_ROI, name_comparison_rev, folders_output_ROI, folders_output_ROI_mask)
    
    
        # Select only materials which are interesting for the comparison 
        if bool_isIncluded :

            # Export peak-ratio map based on ES for multivariate analysis
            if arg_Analysis=='Ratio' and arg_SourceIm=='ES' :
                path_MVA=utils.path_join(outputFd_path,'Stats_ES',name_Short,name_ROI)
                os.makedirs(path_MVA,exist_ok=True)
                
                PeakRatioMaps, edge = load_ES_at_E(path_Dataset_A,row_args['E_A'],row_args['E_B'])
                PeakRatioMap=np.mean(PeakRatioMaps,axis=0)
                PeakRatioMap=np.where(mask_ROI==255,PeakRatioMap,0)
                name_PeakRatioMap=str(row_args['index_MVA'])+'_'+edge+'_'+name_Chems[0]+'_E_'+str(row_args['E_A']).replace('.','_')+'.tif'
                io.imsave(utils.path_join(path_MVA,name_PeakRatioMap,dt='f'),PeakRatioMap.astype(np.float32))
                
            # Export quantification for multivariate analysis
            if arg_Analysis=='Quantification' and arg_SourceIm=='ES' :
                path_MVA=utils.path_join(os.getcwd(),'Stats_ES',name_Short,name_ROI)
                os.makedirs(path_MVA,exist_ok=True)
                
                # If a reference quantification (weight folder) is defined, uses that
                # Note: stopped using it in practice, so always using the stack.
                NoRefQuant=True
                weight_folder=utils.path_join(path_Dataset_A,'weights','weights.tif',dt='f')
                if not os.path.exists(weight_folder) or NoRefQuant :
                    norm_peak_stack, edge = load_ES_at_E(path_Dataset_A,row_args['E_A'],row_args['E_B'])
                    norm_peak=np.mean(norm_peak_stack,axis=0)
                    norm_peak=np.where(mask_ROI==255,norm_peak,np.nan)
                else:
                    _, edge = load_ES_at_E(path_Dataset_A,row_args['E_A'],row_args['E_B'])
                    norm_peak=utils.open_image(weight_folder)
                norm_peak=norm_peak/np.nanmean(norm_peak)
                name_normPeak=str(row_args['index_MVA'])+'_'+edge+'_'+name_Chems[0]+'_E_'+str(row_args['E_A']).replace('.','_')+'.tif'
                io.imsave(utils.path_join(path_MVA,name_normPeak,dt='f'),norm_peak.astype(np.float32))               
            
            # Calculate NNMA-extrema spectra
            path_NNMF=utils.path_join(path_Dataset_A,'NNMF_images')
            path_NNMF_map=utils.path_join(path_NNMF,f'{name_Chems[0]}.tif',dt='f')
            if os.path.exists(path_NNMF_map)  :
                NNMF_map=utils.open_image(path_NNMF_map)
                
                # >> SAVE FOR THE MVA ANALYSIS
                path_MVA=utils.path_join(outputFd_path,'Stats_NNMF',name_Short,name_ROI)
                os.makedirs(path_MVA,exist_ok=True)
                
                name_NNMFmap=str(row_args['index_MVA'])+'_'+edge+'_'+name_Chems[0]+'_E_'+str(row_args['E_A']).replace('.','_')+'.tif'
                io.imsave(utils.path_join(path_MVA,name_NNMFmap,dt='f'),NNMF_map.astype(np.float32))
                
                # >> CALCULATE NNMF-EXTREMA-SPECTRA ON ALL PEEM STACKS + EXPORT ORIGIN
                # > Loop over folders and find the aligned, processed "_undistrdd_processed_test" E-stacks.
                listpath=[]
                listfolder=[]
                for dirpath, dirnames, filenames in os.walk(path_Dataset_A):
                    for dirname in dirnames:
                        # > Take the preprocessed stack
                        if '_undistrdd_processed_test' in dirname:
                            listpath.append(dirpath)
                            listfolder.append(dirname)
                            
                # > Initialize edge variables
                for dirpath, dirname in zip(listpath,listfolder):
                    folder_Project_Dataset_Edge = os.path.basename(dirpath)
                    folder_DxyE = dirname
                    path_Project_Dataset_Edge=utils.path_join(dirpath)
                    path_Project_Dataset_Edge_Estack=utils.path_join(dirpath,folder_DxyE)
                    edge=utils.find_edge(folder_Project_Dataset_Edge,exp='SIM')
           
                    # > Load stack and energy
                    D_xyE=utils.open_sequence(path_Project_Dataset_Edge_Estack)[0]
                    E=utils.load_E_I0(path_Project_Dataset_Edge,processed=True)[0]
                    
                    # > Apply ROI mask and validity criteria to the peak-ratio map
                    NNMF_map_mskd = np.full_like(NNMF_map, np.nan)
                    mask_ROI_valid_vals = (NNMF_map > 0.001) & mask_ROI==1
                    NNMF_map_flat=NNMF_map[mask_ROI_valid_vals]
                    NNMF_map_mskd[mask_ROI_valid_vals]=NNMF_map_flat
                    
                    fig, axs = plt.subplots(figsize=(10, 10))
                    axs.imshow(NNMF_map_mskd, cmap='gray')
                    axs.set_title('Low NNMF Mask')
                    
                    # > Obtain ratio-extrema ROI
                    treshold_peak_ratio=int(10)
                    mask_low_NNMF=(NNMF_map<np.percentile(NNMF_map_flat,treshold_peak_ratio)) & mask_ROI_valid_vals
                    mask_high_NNMF=(NNMF_map>np.percentile(NNMF_map_flat,100-treshold_peak_ratio)) & mask_ROI_valid_vals
                    D_xyE_low_NNMF=np.where(mask_low_NNMF[:, :, np.newaxis],D_xyE,0)
                    D_xyE_ROI=np.where(mask_ROI_valid_vals[:, :, np.newaxis],D_xyE,0)
                    D_xyE_high_NNMF=np.where(mask_high_NNMF[:, :, np.newaxis],D_xyE,0)
                    
                    # > Inline plot of ratio-extrema ROI
                    test = True
                    if test :
                        # Plotting
                        fig, axs = plt.subplots(2, 2, figsize=(10, 10))
                        axs[1, 0].imshow(mask_low_NNMF, cmap='gray')
                        axs[1, 0].set_title('Low NNMF Mask')
                        
                        axs[0, 0].imshow(mask_ROI_valid_vals, cmap='gray')
                        axs[0, 0].set_title('NNMF Masked')
                        
                        axs[0, 1].imshow(mask_high_NNMF, cmap='gray')
                        axs[0, 1].set_title('High NNMF Mask')
                    
                    # > Average over the extrema ROI
                    if utils.is_included(row_args['E_B'],[E[0],E[-1]]) :
                        Eref=row_args['E_B']
                        print(Eref)
                    else:
                        Eref='max'

                    ax, AV_D_xyE_low_NNMF=utils.scatter_mean_stack(E,D_xyE_low_NNMF,'Low',None,plot=test,refPeak_E =Eref)[0:2]
                    ax, AV_D_xyE_ROI=utils.scatter_mean_stack(E,D_xyE_ROI,'Mean',ax,plot=test,refPeak_E=Eref)[0:2]
                    ax, AV_D_xyE_high_NNMF=utils.scatter_mean_stack(E,D_xyE_high_NNMF,'High',ax,plot=test,refPeak_E=Eref)[0:2]
                    
                    np.savetxt(utils.path_join(path_NNMF,'NNMF_'+name_ROI[:3]+'_'+name_Chems[0]+'_'+edge[:2]+name_Sample[:4]+'.csv',dt='f'), np.vstack((E,AV_D_xyE_low_NNMF,AV_D_xyE_ROI,AV_D_xyE_high_NNMF)).T, delimiter=";", header=f'Energy;Lowest {treshold_peak_ratio}% {name_Chems[0]};Mean;Highest {treshold_peak_ratio}% {name_Chems[0]}')

                    if row_args["ExportOrigin"] :
                        # Save ROI spectrum to origin
                        oplt.AddSheetOrigin(os.getcwd(),'OriginPlots.opju',E,[AV_D_xyE_low_NNMF,AV_D_xyE_ROI,AV_D_xyE_high_NNMF],[name_Chems[0],f'Lowest {treshold_peak_ratio}% ','Mean',f'Highest {treshold_peak_ratio}%'],foldername=f'{folder_Origin}/Ratio/spectrum/{name_Chems[0]}',bookname=f'Rt{arg_SourceIm[:1]}--'+name_ROI[:3]+name_Chems[0]+arg_SourceIm,ShNam=name_ROI[:3]+name_Chems[0][:4]+edge[:2]+name_Sample[:4]+'--RtE')
            
            # Calculate image ratio or colocalization
            else :
                # Run comparison of image A and B
                variablesAB = {
                    'path_Project': os.getcwd(),                         # Project path.
                    'Source_Image': arg_SourceIm,                     # ES or 2E
                    'name_Sample': name_Sample,                        # Ex: Uncycled
                    'name_Short': name_Short,                     # Ex: Uncy
                    'name_ROI': name_ROI,                                   # Ex: NCM
                    'PeakNorm': row_args['E_B'],        # Ex: 851.3 
                    'name_comparison': name_comparison,           # Ex: Nired_011_014
                    'labels': labels,                             # Ex: [011,014]
                    'name_Chem': name_Chems[0],                   # Ex: Ni_red
                    'Analysis': arg_Analysis,             # Ex: Ratio or co-localisation
                    'ExportOrigin': row_args["ExportOrigin"],     # Ex: True
                    'folder_Origin': folder_Origin,
                    'index_Column': row_args["index_Column"],            # Ex: 0
                    'IJ': row_args["ExportIJ"]                   # Ex: False
                }
                
                # Run comparison of image B and A
                if row_args["Process"] :
                    df=PEEM_2i_comp2(df, images_A,images_B,mask_ROI, variablesAB)

                if row_args["Process_rev"] :
                    variablesBA = {
                        'path_Project': os.getcwd(),
                        'Source_Image': arg_SourceIm,
                        'name_Sample': name_Sample,
                        'name_Short': name_Short,
                        'name_ROI': name_ROI,
                        'PeakNorm': row_args['E_B'],
                        'name_comparison': name_comparison_rev,
                        'labels': labels[::-1],
                        'name_Chem': name_Chems[1],
                        'Analysis': arg_Analysis,
                        'ExportOrigin': row_args["ExportOrigin"],
                        'index_Column': row_args["index_Column"],
                        'IJ': row_args["ExportIJ"],
                    }
                    _=PEEM_2i_comp2(df, images_B,images_A,mask_ROI, variablesBA)
        
    # >> EXPORT THE SUMMARY SHEET
    if row_args["ExportOrigin"] and not row_args['Analysis']=='Quantification' :
        # Save the dataframe to origin
        loc=os.getcwd()
        file='OriginPlots.opju'
        oplt.AddDFOrigin(loc, file,df, bookname=f'complist{arg_SourceIm}',foldername='2E_energies',ShNam=folder_Output)
    
    if row_args["Process"] and not arg_Analysis=='Quantification' :
        update_excel_sheet(df,path_comparisons_list,folder_Output)
    
    end_time = time.time()
    print(f"Execution time: {end_time - start_time} seconds")

def PEEM_2i_comp2(df, images_A, images_B, mask_ROI, params):
    """
    Process and compare two sets of images.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to store results.
    images_A : np.ndarray
        First set of images.
    images_B : np.ndarray
        Second set of images.
    mask : np.ndarray or similar
        Mask to apply to the images.
    params : dict
        Dictionary containing the 'name_Sample', 'name_ROI', 'name_Chem', 'comparison_label', 'labels', and fields.

    Returns
    -------
    df : pd.DataFrame
        DataFrame with the results.
    """
    # >> Extract the parameters for this comparison

    arg_SourceIm = params['Source_Image']
    name_Sample = params['name_Sample']
    sampleFd_path = utils.path_join(os.getcwd(),name_Sample)
    outputFd_path = utils.path_join(os.getcwd(),'_Output')
    outputFd2E_path = utils.path_join(outputFd_path, name_Sample, f'{arg_SourceIm}')
    name_Short = params['name_Short']
    name_ROI = params['name_ROI']
    complabel = params['name_comparison']
    name_Chem = params['name_Chem']
    Analysis = params['Analysis']
    
    labels = params['labels']

    ExportOrigin = params["ExportOrigin"]
    folder_Origin = params["folder_Origin"]
    index_Column = params["index_Column"]
    E_peak_B=params["PeakNorm"]

    # >> CHECKS ON TWO-IMAGES AND AVERAGING
    # Type
    if not isinstance(images_A, np.ndarray) or not isinstance(images_B, np.ndarray):
        raise TypeError('Both stackA and stackB must be numpy arrays')
    if images_A.dtype == np.uint8 or images_A.dtype == np.uint16: images_A=images_A.astype(np.float32)
    if images_B.dtype == np.uint8 or images_B.dtype == np.uint16: images_B=images_B.astype(np.float32)

    # > Add an extra dimension if stackA and stackB are single images
    images_A = np.expand_dims(images_A, axis=2) if len(images_B.shape) == 2 else images_A
    images_B = np.expand_dims(images_B, axis=2) if len(images_B.shape) == 2 else images_B
    
    # > Assert that the dimensions are the same for both images
    (pA,nA,mA)=images_A.shape
    (pB,nB,mB)=images_B.shape
    assert nA == nB and mA == mB
    if not (pA == pB) :
        p = min(pA,pB)
        warnings.warn("The numer of replicates for image A and B are not the same, using only f{p} images")        
        images_A = images_A[0:p,:,:]
        images_B = images_B[0:p,:,:]
        
    # > Calculate the average 
    imageA=np.nanmean(images_A,axis=0)
    imageB=np.nanmean(images_B,axis=0)
    
    # >> MASKS
    if not isinstance(mask_ROI,np.ndarray) :
        [mskimageA,mskimageB],mask_ROI=utils.image_well_defined([imageA,imageB],axis=0)
    else:
        [mskimageA,mskimageB],mask_ROI=utils.image_well_defined([imageA*mask_ROI,imageB*mask_ROI],axis=0)
    blmsk_ROI=mask_ROI==1 # Boolean mask of the region of interest
    
    if Analysis =='Ratio' :

        # >> CALCULATE PEAK-RATIO MAP AND OTHER CONTRASTS (PEAK-DIFFERENCE MAP, EQUALISED PEAK-RATIO MAP, SSIM MAP, MSE MAP)
        print(f'Current execution: peak-ratio map of {complabel} at {name_ROI}.')
        image_dict, abbr_to_label_im, value_dict, abbr_to_label_val = PEEM_peak_ratio(imageA, imageB, msksegm = mask_ROI)
        
        # >> EXPORT PEAK-RATIO MAP TO ORIGIN
        if ExportOrigin :
            #Export to origin as grayscale (Gs) 
            oplt.AddImageOrigin(image_dict['div_A_B'], f'{folder_Origin}/Ratio/images/{name_Chem}', 'G'+arg_SourceIm+name_Chem, 'G'+name_Short+name_Chem+name_ROI+'ugraphGs')
    
        # >> EXPORT COMPUTED IMAGES TO FILES
        # > Initialise base directory
        path_output_ROI = utils.path_join(outputFd2E_path,f'{arg_SourceIm}_2E_ratio', name_ROI, complabel)
        abbimAB = list(image_dict.keys())
        namelistAB=[f'{abbr}_{complabel}_{name_ROI}' for abbr in abbimAB]
        imagesAB = list(image_dict.values())
        image_labelsAB = [abbr_to_label_im[abbr] for abbr in abbimAB]
    
        # > Save the raw and calculated images
        for img, img_abbr, img_name, img_label in zip(imagesAB,abbimAB,namelistAB,image_labelsAB):
            img = np.where(img == 0, np.nan, img)

            # Save raw comparison image
            path_output_ROI_raw = utils.path_join(path_output_ROI, "raw",f'{img_name}.tif',dt='f')
            io.imsave(path_output_ROI_raw, img.astype(np.float32))

            # Contrast correction (CC)
            path_output_ROI_gray_cc = path_output_ROI_raw.replace("raw", 'grayscale_cc')
            image_adjust_contrast(path_output_ROI_raw, path_output_ROI_gray_cc)
     
            # Calculate and save mask for various treshold
            image_mask_save(path_output_ROI_gray_cc,img,blmsk_ROI,trsh=10)
            image_mask_save(path_output_ROI_gray_cc,img,blmsk_ROI,trsh=5)
            image_mask_save(path_output_ROI_gray_cc,img,blmsk_ROI,trsh=2)
            image_mask_save(path_output_ROI_gray_cc,img,blmsk_ROI,trsh=1)
            
            # Add scalebar and colormap in ImageJ to grayscale images
            if params['IJ']:
                IJ_col_and_scalebar(path_output_ROI_gray_cc, utils.path_join(path_output_ROI, "color"))
            
        # >> CALCULATE RATIO-EXTREMA-SPECTRA ON ALL PEEM STACKS + EXPORT ORIGIN
        # > Loop over folders and find the aligned, processed "_undistrdd_processed_test" E-stacks.
        listpath=[]
        listfolder=[]
        for dirpath, dirnames, filenames in os.walk(sampleFd_path):
            for dirname in dirnames:
                # > Take the preprocessed stack
                if '_undistrdd_processed_test' in dirname:
                    listpath.append(dirpath)
                    listfolder.append(dirname)
                    
        # > Initialize edge variables
        for dirpath, dirname in zip(listpath,listfolder):
            edgeFd_name = os.path.basename(dirpath)
            edge=utils.find_edge(edgeFd_name,exp='SIM')
            edgeFd_path=utils.path_join(sampleFd_path,edgeFd_name)
            stackFd_name = dirname
            path_Project_Dataset_Edge_Estack=utils.path_join(edgeFd_path,stackFd_name)
   
            # > Load stack and energy
            D_xyE=utils.open_sequence(path_Project_Dataset_Edge_Estack)[0]
            E=utils.load_E_I0(edgeFd_path,processed=True)[0]
            
            # > Apply ROI mask and validity criteria to the peak-ratio map
            peak_ratio_map=image_dict["div_A_B"]
            ratio_mskd = np.full_like(peak_ratio_map, np.nan)
            mask_ROI_valid_vals = (peak_ratio_map > 0.01) & blmsk_ROI & (peak_ratio_map < 30)
            peak_ratio_map_flat=peak_ratio_map[mask_ROI_valid_vals]
            ratio_mskd[mask_ROI_valid_vals]=peak_ratio_map_flat
            
            # > Obtain ratio-extrema ROI
            treshold_peak_ratio=int(10)
            mask_low_ratio=(ratio_mskd<np.percentile(peak_ratio_map_flat,treshold_peak_ratio))
            mask_high_ratio=ratio_mskd>np.percentile(peak_ratio_map_flat,100-treshold_peak_ratio)
            D_xyE_low_ratio=np.where(mask_low_ratio[:, :, np.newaxis],D_xyE,0)
            D_xyE_ROI=np.where(mask_ROI_valid_vals[:, :, np.newaxis],D_xyE,0)
            D_xyE_high_ratio=np.where(mask_high_ratio[:, :, np.newaxis],D_xyE,0)
            
            # > Inline plot of ratio-extrema ROI
            test = True
            if test :
                # Plotting
                fig, axs = plt.subplots(2, 2, figsize=(10, 10))
                
                axs[0, 0].imshow(mask_ROI_valid_vals, cmap='gray')
                axs[0, 0].set_title('Ratio Masked')
                
                axs[0, 1].imshow(mask_high_ratio, cmap='gray')
                axs[0, 1].set_title('High Mask')
                
                axs[1, 0].imshow(mask_low_ratio, cmap='gray')
                axs[1, 0].set_title('Low Mask')
            
            # > Average over the extrema ROI
            if utils.is_included(E_peak_B,[E[0],E[-1]]) :
                Eref=params["PeakNorm"]
                print(Eref)
            else:
                Eref='max'

            ax, AV_D_xyE_low_ratio=utils.scatter_mean_stack(E,D_xyE_low_ratio,'Low',None,plot=test,refPeak_E =Eref)[0:2]
            ax, AV_D_xyE_ROI=utils.scatter_mean_stack(E,D_xyE_ROI,'Mean',ax,plot=test,refPeak_E =Eref)[0:2]
            ax, AV_D_xyE_high_ratio=utils.scatter_mean_stack(E,D_xyE_high_ratio,'High',ax,plot=test,refPeak_E =Eref)[0:2]
            
            np.savetxt(utils.path_join(path_output_ROI,arg_SourceIm+'_'+name_ROI[:3]+'_'+name_Chem+'_'+edge[:2]+name_Sample[:4]+'.csv',dt='f'), np.vstack((E,AV_D_xyE_high_ratio,AV_D_xyE_low_ratio)).T, delimiter=";", header=f'Energy;Highest {treshold_peak_ratio}% {name_Chem}, Lowest {treshold_peak_ratio}% {name_Chem}')

            if ExportOrigin :
                # Save ROI spectrum to origin
                oplt.AddSheetOrigin(os.getcwd(),'OriginPlots.opju',E,[AV_D_xyE_low_ratio,AV_D_xyE_ROI,AV_D_xyE_high_ratio],[name_Chem,f'Lowest {treshold_peak_ratio}% ','Mean',f'Highest {treshold_peak_ratio}%'],foldername=f'{folder_Origin}/Ratio/spectrum/{name_Chem}',bookname=f'Rt{arg_SourceIm[:1]}--'+name_ROI[:3]+name_Chem+arg_SourceIm,ShNam=name_ROI[:3]+name_Chem[:4]+edge[:2]+name_Sample[:4]+'--RtE')
 
    elif Analysis =='Co-localisation' :
        im_dict, value_dict, abbr_to_label_val, export_corr = PEEM_colocalisation(imageA, imageB, msksegm = mask_ROI)
        
        # >> Histogramm comp
        if ExportOrigin:
            binsA, binsB = im_dict['BinsA']*100, im_dict['BinsB']*100
            HistA, HistB = im_dict['HistA']*100, im_dict['HistB']*100
            oplt.AddSheetOrigin(os.getcwd(),'OriginPlots.opju', binsA, HistA,['',labels[0]],foldername=f'{folder_Origin}/coloc/hist/{name_Chem}',shiftCol=0,bookname='G'+arg_SourceIm+'--'+name_ROI+name_Chem+'hist',ShNam=name_ROI+name_Chem+name_Short+''+'hist')
            oplt.AddSheetOrigin(os.getcwd(),'OriginPlots.opju', binsB, HistB,['',labels[1]],foldername=f'{folder_Origin}/coloc/hist/{name_Chem}',shiftCol=1,bookname='G'+arg_SourceIm+'--'+name_ROI+name_Chem+'hist',ShNam=name_ROI+name_Chem+name_Short+''+'hist')

        # >> CORRELATION TO ORIGIN
        if ExportOrigin :
            XPEEM_2i_exportCorrOrigin(folder_Origin, arg_SourceIm, name_ROI, name_Short, name_Chem, export_corr)

        # >> BUILD OVERLAY IMAGE A, B
        # > Stack image in the overlay and save
        pathComp = utils.path_join(outputFd2E_path,f'{arg_SourceIm}_2E_ratio', name_ROI, complabel,'Overlay')
        rkA, rkB = im_dict['RkA'], im_dict['RkB']

        nameOverlay=f'OverlayAB_{complabel}_{name_ROI}.tif'
        blue_channel = rkB.astype(np.uint8)        
        green_channel = np.zeros_like(rkA, dtype=np.uint8)
        red_channel = rkA.astype(np.uint8)
        
        BGRimage = np.stack((blue_channel, green_channel, red_channel), axis=-1)
        cv2.imwrite(utils.path_join(pathComp,nameOverlay,dt='f'), BGRimage)
        
        # >> CALCULATE HUE TRESHOLDS
        def mask_hue(Im, file_name, comp_name):
            """ Function calculating the hue treshold-based masks """
            # Save extremum image
            io.imsave(utils.path_join(pathComp, f'{file_name}_{complabel}_{name_ROI}.tif', dt='f'), Im)
            # Save mask
            for trsh in [10, 5, 2, 1]:
                name = f'{comp_name}_{complabel}_{name_ROI}_masks_{trsh}.tif'
                mask=image_mask_save(utils.path_join(pathComp, name, dt='f'), Im, blmsk_ROI, trsh=trsh, comp=True)
            return mask
        
        # > Find hue mask for each color
        redzones,maskAred = image_threshold_hue(BGRimage, 223, 255)
        diffAB = redzones[:,:,2]-redzones[:,:,1]
        redmask = mask_hue(diffAB, 'Red_diff_A_B', 'CompMaskA')
        magentazones, maskABmagenta = image_threshold_hue(BGRimage, 200, 223)
        sumAB = magentazones[:,:,1] + magentazones[:,:,2]
        magentamask = mask_hue(sumAB, 'Magenta_summ_A_B', 'CompMaskAB')
        bluezones,maskBblue = image_threshold_hue(BGRimage, 170, 200)
        diffBA = bluezones[:,:,1]-bluezones[:,:,2]
        bluemask = mask_hue(diffBA, 'Blue_diff_B_A', 'CompMaskB')
        
        # > Export Hue fractions to Origin        
        if ExportOrigin :
            loc=os.getcwd()
            nrpxl=np.count_nonzero(maskAred+maskABmagenta+maskBblue)
            if nrpxl==0:
                nrpxl=1
            oplt.AddSheetOrigin(loc,'OriginPlots.opju',np.array([[name_Sample]]), np.array([[np.count_nonzero(maskAred)/nrpxl,np.count_nonzero(maskABmagenta)/nrpxl,np.count_nonzero(maskBblue)/nrpxl]]),['Sample','Red','Magenta','Blue'],foldername=f'{folder_Origin}/coloc/metrics/{name_Chem}',shiftRow=index_Column,bookname=arg_SourceIm+'--'+name_ROI+name_Chem+'Metr',ShNam=name_ROI+name_Chem+''+'Over--lapMetr3',typedata='row')
            
        # >> APPLY HUE MASKS ON ALL PEEM STACKS + EXPORT ORIGIN
        listpath=[]
        listfolder=[]
        for dirpath, dirnames, filenames in os.walk(sampleFd_path):
            for dirname in dirnames:
                # > Take the preprocessed stack
                if '_undistrdd_processed_test' in dirname:
                    listpath.append(dirpath)
                    listfolder.append(dirname)
                    
        # > Initialize edge variables
        for dirpath, dirname in zip(listpath,listfolder):
            
            edgeFd_name = os.path.basename(dirpath)
            edge=utils.find_edge(edgeFd_name,exp='SIM')
            edgeFd_path=utils.path_join(sampleFd_path,edgeFd_name)
            
            stackFd_name = dirname
            stackFd_path=utils.path_join(edgeFd_path,stackFd_name)
            
            # > Load stack and energy
            D_xyE=utils.open_sequence(stackFd_path)[0]
            E=utils.load_E_I0(edgeFd_path,processed=True)[0]
            
            # > Multiply with mask and convert to ndarray
            D_xyE_red=np.where(redmask[:, :,np.newaxis],D_xyE,0)
            D_xyE_mag=np.where(magentamask[:, :,np.newaxis],D_xyE,0)
            D_xyE_blue=np.where(bluemask[:, :,np.newaxis],D_xyE,0)
            
            # > Average
            test=False
            ax, AV_D_xyE_red=utils.scatter_mean_stack(E,D_xyE_red,'Red',None,plot=test,refPeak_E=Eref)[0:2]
            ax, AV_D_xyE_mag=utils.scatter_mean_stack(E,D_xyE_mag,'Magenta',ax,plot=test,refPeak_E=Eref)[0:2]
            ax, AV_D_xyE_blue=utils.scatter_mean_stack(E,D_xyE_blue,'Blue',ax,plot=test,refPeak_E=Eref)[0:2]
            
            dircomp=utils.path_join(outputFd2E_path,f'{arg_SourceIm}_2E_coloc', name_ROI, complabel)
            np.savetxt(utils.path_join(dircomp,arg_SourceIm+'_'+name_ROI[:3]+'_'+name_Chem+'_'+edge[:2]+name_Sample[:4]+'.csv',dt='f'), np.vstack((E,[AV_D_xyE_red,AV_D_xyE_mag,AV_D_xyE_blue])).T, delimiter=";", header='Energy;Red;Magenta;Blue')
            
            if ExportOrigin :
                # Spectrum sheet in origin
                # Save ROI spectrum to origin
                oplt.AddSheetOrigin(os.getcwd(),'OriginPlots.opju',E,[AV_D_xyE_red,AV_D_xyE_mag,AV_D_xyE_blue],[name_Chem,'Red','Magenta','Blue'],foldername=f'{folder_Origin}/Coloc/spectrum/{name_Chem}',bookname=f'{arg_SourceIm}RB--'+name_ROI[:3]+name_Chem[:4]+edge[:2],ShNam=name_ROI[:3]+name_Chem[:4]+edge[:2]+name_Sample+'--RFB')

        # > Comparative plot of the SCC
        if ExportOrigin :
            loc=os.getcwd()
            oplt.AddSheetOrigin(loc,'OriginPlots.opju',np.array([[name_Sample]]), np.array([[value_dict['SCC'],value_dict['SCI']]]),[name_Chem]*3,foldername=f'{folder_Origin}/Coloc/metrics/{name_Chem}',shiftRow=index_Column,bookname=arg_SourceIm+'--'+name_ROI+name_Chem+'Metr',ShNam=name_ROI+name_Chem+''+'SCC--Metr1',typedata='row',colNames=['Sample','SCC','SCI'])
            
    # >> UPDATE SUMMARY DATASHEET
    col = pd.DataFrame([{
        'name_Sample': name_Sample,
        'name_ROI': name_ROI,
        'name_Comparison': complabel
    }])
    value_labelsAB = [abbr_to_label_val[abbr] for abbr in list(value_dict.keys())]
    for val, label in zip(list(value_dict.values()), value_labelsAB):
        col[label] = val
    
    df = update_dataframe(df, col)

    return df

def test_PEEM_compare2i():
    # Create two random 3D numpy arrays (i.e., image stacks)
    stackA = np.random.rand(5, 256, 256)
    stackB = np.random.rand(5, 256, 256)

    # Compare the two stacks and plot the results
    value_dict = compare_and_plot(stackA, stackB, "Random image stacks comparison")

    # Print the calculated values with their labels
    df = pd.DataFrame({'Values': value_dict})

    print(df)

def test_PEEM_compare2i_2(mean1, mean2, size=100, sigma=10, num_images=5):
    # Create empty image stacks
    stackA = np.zeros((num_images, size, size))
    stackB = np.zeros((num_images, size, size))

    # Create a grid of (x, y) coordinates
    x, y = np.meshgrid(np.linspace(0, size-1, size), np.linspace(0, size-1, size))

    # For each image in the stack
    for i in range(num_images):
        # Create a 2D Gaussian distribution
        d = np.sqrt((x-mean1[0])**2 + (y-mean1[1])**2)
        stackA[i] = np.exp(-(d**2 / (2.0 * sigma**2)))
        
        d = np.sqrt((x-mean2[0])**2 + (y-mean2[1])**2)
        stackB[i] = np.exp(-(d**2 / (2.0 * sigma**2)))

    # Compare the two Gaussian images and plot the results
    value_dict = compare_and_plot(stackA, stackB, "Gaussian images comparison")

    # Print the calculated values with their labels
    df = pd.DataFrame({'Values': value_dict})

    print(df)
    
def test_PEEM_compare2i_3(typeIm = 'chemcomp_highCorr'):
    # Comparison 1: SO4 and layered oxide (low correlation)
    # The image was pre-processed by dividing by the pre-edge 
    if typeIm == 'chemcomp_lowCorr':
        path= r'Test/PEEM/compare2i'
        stackA = np.array(Image.open(utils.path_join(path,'normA_SuLa_SO4_036_NCM.tif',dt='f')))
        stackB = np.array(Image.open(utils.path_join(path,'normB_SuLa_SO4_036_NCM.tif',dt='f')))
    # Comparison 2: oxidized Co and layered oxide (High correlation)
    # The image was pre-processed by dividing by the pre-edge.
    if typeIm == 'chemcomp_highCorr':
        path= r'Test/PEEM/compare2i'
        stackA = np.array(Image.open(utils.path_join(path,'normA_CooLa_024_036_NCM.tif',dt='f')))
        stackB = np.array(Image.open(utils.path_join(path,'normB_CooLa_024_036_NCM.tif',dt='f')))
    elif typeIm == 'peak':
        path= r'D:\Documents\a PSI\Data\Data analysis\spyder\2107_Progress_work\0410_SIM beamline 1\PEEM_py\Uncycled\Ni_uncycled\2_energies'
        stackA = np.array(Image.open(utils.path_join(path,'11_DIV_09_on_10.tif',dt='f')))
        stackB = np.array(Image.open(utils.path_join(path,'14_DIV_12_on_13.tif',dt='f')))
        
        stackA,stackB=utils.images_outliers2nan(stackA,stackB)

    # Compare the two stacks and plot the results
    value_dict = compare_and_plot(stackA, stackB, "Stacks comparison")

    # Print the calculated values with their labels
    df = pd.DataFrame({'Values': value_dict})

    print(df)
    
def test_PEEM_compare2i_4():
    # Generate images with similar histograms
    stackA = np.random.normal(loc=0, scale=1, size=(5, 256, 256))
    stackB = np.random.normal(loc=0, scale=1, size=(5, 256, 256))
    similar_values = compare_and_plot(stackA, stackB, "Stacks with similar histograms")
    

    # Generate images with slightly overlapping histograms
    stackA = np.random.normal(loc=0, scale=1, size=(5, 256, 256))
    stackB = np.random.normal(loc=1, scale=1, size=(5, 256, 256))
    slight_overlap_values = compare_and_plot(stackA, stackB, "Stacks with slightly overlapping histograms")

    # Generate images with non-overlapping histograms
    stackA = np.random.normal(loc=0, scale=1, size=(5, 256, 256))
    stackB = np.random.normal(loc=5, scale=1, size=(5, 256, 256))
    non_overlap_values = compare_and_plot(stackA, stackB, "Stacks with non-overlapping histograms")

    # Create a DataFrame to summarize the results
    df = pd.DataFrame({
        'Similar Histograms': similar_values,
        'Slight Overlap': slight_overlap_values,
        'Non-overlapping': non_overlap_values
    })

    print(df)    

def compare_and_plot(stackA, stackB, title):
    # Compare the two stacks
    image_dict, abbr_to_label_im, value_dict, abbr_to_label_val, export_corr = PEEM_colocalisation(stackA, stackB, test = True)

    # Determine the number of rows for the subplot grid
    nrows = int(np.ceil(len(image_dict) / 3))

    # Plot the resulting images
    fig, axs = plt.subplots(nrows, 3, figsize=(15, 5 * nrows))
    axs = axs.flatten()  # Flatten the array of axes to make it easier to iterate over
    for i, (label, img) in enumerate(image_dict.items()):
        axs[i].imshow(img, cmap='gray')
        axs[i].set_title(label)

    # Plot the histograms
    bins = np.linspace(-5, 5, 100)
    axs[-1].hist(stackA.flatten(), bins, alpha=0.5, label='Stack A')
    axs[-1].hist(stackB.flatten(), bins, alpha=0.5, label='Stack B')
    axs[-1].set_title('Histograms')
    axs[-1].legend(loc='upper right')

    plt.tight_layout()
    plt.suptitle(title)
    plt.show()

    # Return the calculated values
    return value_dict

def update_dataframe(df: pd.DataFrame, col: pd.DataFrame) -> pd.DataFrame:
    """
    Update a DataFrame based on the values in two lists and another DataFrame.

    Parameters:
    df (pd.DataFrame): The DataFrame to update.
    col (pd.DataFrame): A DataFrame containing additional data to add to df.

    Returns:
    pd.DataFrame: The updated DataFrame.
    """
    if col.empty:
        raise ValueError("The col argument should contain values")
    assert all(isinstance(x, (float, str)) for x in col.values.flatten()), "DataFrame contains non-float and non-string values"

    # Reset the index of df before the comparison
    df = df.reset_index(drop=True)

    # # Check that the two dataframes have the same column name
    if len(df.columns.tolist()) > len(col.columns.tolist()):
        raise ValueError("There should always have more column in the updating dataframe.")
    elif len(df.columns.tolist()) < len(col.columns.tolist()):
        warnings.warn("There is more column in the coll dataframe than was defined in the sheet summary, the new column will be added. \n Please delete Summary if you want to reset the columns after a modification of the code.")
        
    # Create a mask for the rows that match the conditions
    if df.empty:
        df = col
        print('The sheet Summary was not found.')
        print('A new sheet Summary will be created in the comparison list. ')
    # If the data does not already exist in the DataFrame, append it
    else:
        exist_line = (df['name_Sample'] == col['name_Sample'][0]) & (df['name_ROI'] == col['name_ROI'][0]) & (df['name_Comparison'] == col['name_Comparison'][0])
        if not exist_line.any():
            df = pd.concat([df, col], ignore_index=True)
        # If the data exists in the DataFrame, update it.
        else:
            for column in col.columns:
                if column not in df.columns:
                    df[column] = np.nan  # or some default value
            # Get the common columns
            common_cols = df.columns.intersection(col.columns)
            
            # Assign values only to the common columns
            df.loc[exist_line, common_cols] = col[common_cols].values[0]
    
    return df

def test_update_dataframe():
    # Test case 1: Empty DataFrame
    df = pd.DataFrame(columns=['name_Sample', 'name_ROI', 'name_Comparison'])
    col = pd.DataFrame({'name_Sample': ['sample1'], 'name_ROI': ['material1'], 'name_Comparison': ['label1'], 'Kind': ['kind1'], 'label1': ['value1'], 'label2': ['value2']})
    
    with warnings.catch_warnings(record=True) as w:
        result = update_dataframe(df, col)
        assert result.equals(col)
        assert len(w) == 1
    
    # Test case 2: Empty col
    df = pd.DataFrame({'Sample': ['sample1'], 'Material': ['material1'], 'Labels': ['label1'], 'label1': ['value1'], 'label2': ['value2']})
    col = pd.DataFrame(columns=['Sample', 'Material', 'Labels', 'Kind', 'label1', 'label2'])    
    with pytest.raises(ValueError):
        result = update_dataframe(df, col)
    
    # Test case 3: More columns in df
    df = pd.DataFrame({'Sample': ['sample1'], 'Material': ['material1'], 'Labels': ['label1'], 'label1': ['value1'], 'label2': ['value2'], 'label3': ['value3']})
    col = pd.DataFrame({'Sample': ['sample1'], 'Material': ['material1'], 'Labels': ['label1'], 'label1': ['value1'], 'label2': ['value2']})
    with pytest.raises(ValueError):
        update_dataframe(df, col)
        
    # Test case 4: More columns in col
    df = pd.DataFrame({'Sample': ['sample1'], 'Material': ['material1'], 'Labels': ['label1'], 'label1': ['value1'], 'label2': ['value2']})
    col = pd.DataFrame({'Sample': ['sample1'], 'Material': ['material1'], 'Labels': ['label1'], 'label1': ['value1'], 'label2': ['value2'], 'label3': ['value3']})
    with warnings.catch_warnings(record=True) as w:
        # Cause all warnings to always be triggered.
        warnings.simplefilter("always")
        # Trigger a warning.
    
        result = update_dataframe(df, col)
        assert result.shape[1] == col.shape[1]
        assert len(w) == 2

def PEEM_peak_ratio(imageA: np.ndarray, imageB: np.ndarray, msksegm: np.ndarray = None, test: bool = False) -> tuple[list[np.ndarray], list[str], list[float], list[str]]:
    """
    Compares two image stacks by calculating the differences, ratios, SSIM, MSE, Pearson correlation, and histogram comparisons between them.

    Args:
        stackA (ndarray): The first image stack.
        stackB (ndarray): The second image stack.
        msksegm (ndarray): Mask to apply.

    Returns:
        tuple: A tuple containing four lists:
            - The first list contains the raw images and the calculated images (difference, division, SSIM image, MSE image).
            - The second list contains the labels for the images in the first list.
            - The third list contains the calculated values (SSIM value, MSE value, correlation, histogram correlation, chi square, intersection, Bhattacharyya).
            - The fourth list contains the labels for the values in the third list.
    """    


    # >> Initialize variables.
    # > Global mask for the stack
    boolmsk=msksegm==1
    # > Dictionary to save the result image
    avg_images = {
        'RawA': imageA, 'RawB': imageB, 'normA': [], 'normB': [], 'rkA': [], 'rkB': [], 'diff_A_B': [], 'div_A_B': [], 'div_A_B_processed': [], 'div_B_A': [], 'div_A_B_rk': [], 'SSIM': []        
    }
    # > Dictionary to save the result values for each pair of image
    metrics = {'SSIM': None, 'MSE': None, 'Mean': None, 'Median': None, 'Mode':None, 'IQR': None}

    # >> QUANTITATIVE IMAGE COMPARISON
    # > Rank
    rkA, rkB = rank_pixels(imageA, boolmsk), rank_pixels(imageB, mask=boolmsk)
    normrkA, normrkB = utils.image_positivehist(rkA, normalize='8bit', mask = msksegm)[0], utils.image_positivehist(rkB, normalize='8bit', mask = msksegm)[0]
    avg_images['rkA']=normrkA
    avg_images['rkB']=normrkB
    
    # > Difference and ratio
    diffAB=np.where(boolmsk,imageA - imageB,0)
    avg_images['diff_A_B']=diffAB
    divAB=np.where(boolmsk,imageA / imageB, 0)
    avg_images['div_A_B']=divAB
    avg_images['div_A_B_processed']=divAB
    divBA=np.where(boolmsk,imageB / imageA, 0)
    avg_images['div_B_A']=divBA
    avg_images['div_A_B_rk']=utils.image_positivehist(rank_pixels(divAB), normalize='8bit', mask = msksegm)[0]    
    
    # > Normalize image to range 0-255
    normAk, normBk = utils.image_positivehist(imageA, normalize='8bit', mask = msksegm)[0], utils.image_positivehist(imageB, normalize='8bit', mask = msksegm)[0]
    avg_images['normA']=normAk
    avg_images['normB']=normBk
    
    # > Calculate SSIM and MSE for each image
    _, ssim_imgk = ssim(imageA, imageB, data_range=255, multichannel=False, full=True)
    metrics['SSIM']=np.mean(ssim_imgk[boolmsk])
    avg_images['SSIM']=np.where(boolmsk,ssim_imgk, 0)
    flnorA, flnorB = imageA[msksegm], imageB[msksegm]
    metrics['MSE']=mean_squared_error(flnorA, flnorB)


    # >> DESCRIPTIVE STATISTICS ON IMAGE RATIO'S HISTOGRAMM
    # Compute the histogram
    datarange=np.max(utils.median_filter3(divAB,nr_pixel=3))*2
    hist = cv2.calcHist([divAB.astype(np.float32)], [0], msksegm, [1000], [0, datarange])
    
    # Normalize the histogram to get a "density"
    hist /= np.sum(hist)
    
    # Compute mean, median, mode, interquartile range (10-90%) and skewness
    flatratio=divAB[boolmsk & (divAB > 0)]
    metrics['Mean']= np.mean(flatratio)
    metrics['Median'] = np.median(flatratio)
    med10=metrics['Median']*10
    avg_images['div_A_B'][divAB>med10]=med10
    avg_images['div_B_A'][divBA>1/med10]=1/med10
    metrics['Mode'] = stats.mode(flatratio, axis=None)[0]
    metrics['IQR'] = np.percentile(flatratio, 90) - np.percentile(flatratio, 10)
    
    # > Create dictionary to store the result images and return them to the calling function
    image_labels = ['Raw Image A', 'Raw Image B', 'Normalized image A', 'Normalized image B', 'Ranked image A', 'Ranked image B', 'Difference', 'Ratio', 'Ratio processed', 'Ratio Reversed', 'Ranked ratio', 'Structural Similarity Index Measure (SSIM) Image']
    abbr_to_label_im = {abbr: label for abbr, label in zip(avg_images.keys(), image_labels)}
    
    metrics_labels = ['Structural Similarity Index Measure', 'Mean Square Error', 'Mean ratio', 'Median ratio', 'Mode ratio', 'Ratio inter quartile range 10-90 ']
    abbr_to_label_val = {abbr: label for abbr, label in zip(metrics.keys(), metrics_labels)}

    return avg_images, abbr_to_label_im, metrics, abbr_to_label_val

def PEEM_colocalisation(imageA: np.ndarray, imageB: np.ndarray, msksegm: np.ndarray = None, test: bool = False) -> tuple[list[np.ndarray], list[str], list[float], list[str]]:
    """
    Compares two image stacks by comparing their histogram and doing colocalisation analysis on the images.

    Args:
        imageA (ndarray): The first image.
        imageA (ndarray): The second image.
        msksegm (ndarray): Mask to apply.

    Returns:
        tuple: A tuple containing four lists:
            - The first list contains the raw images and the calculated images (difference, division, SSIM image, MSE image).
            - The second list contains the labels for the images in the first list.
            - The third list contains the calculated values (SSIM value, MSE value, correlation, histogram correlation, chi square, intersection, Bhattacharyya).
            - The fourth list contains the labels for the values in the third list.
    """    
    boolmsk=msksegm==1
    im_dict = {
        'RkA': None, 'RkB': None, 'HistA': None, 'HistB': None, 'BinsA': None, 'BinsB': None}
    
    def get_hist(im):
        nr_bins=1000
        flatim=im[boolmsk]
        q98 = np.percentile(flatim, 98)
        q02 = np.percentile(flatim, 2)
        hist = cv2.calcHist([flatim.astype(np.float32)], [0], None, [nr_bins], [q02, q98])
    
        # Normalize the histogram to get a "density"
        hist /= np.sum(hist)
        bins=np.arange(0,1,1/nr_bins)
    
        return hist, bins
    
    # >> HISTOGRAMM RAW RATIOS
    # > Calculate the histograms (exclude the background values)
    im_dict['HistA'],im_dict['BinsA'] = histA, _ = get_hist(imageA)
    im_dict['HistB'],im_dict['BinsB'] = histB, _ = get_hist(imageB)

    # > Compare the histograms
    hist_dict = {
        'HistCorr': [], 'HChiSq': [], 'HInter': [], 'Bhatt': []}
    hist_dict['HistCorr']=cv2.compareHist(histA, histB, cv2.HISTCMP_CORREL)
    hist_dict['HChiSq']=cv2.compareHist(histA, histB, cv2.HISTCMP_CHISQR)
    hist_dict['HInter']=cv2.compareHist(histA, histB, cv2.HISTCMP_INTERSECT)
    hist_dict['Bhatt']=cv2.compareHist(histA, histB, cv2.HISTCMP_BHATTACHARYYA)
    hist_labels = ['Histogram Correlation', 'Chi Square', 'Intersection', 'Bhattacharyya']    
    
    # EQUALIZED RATIOS INTENSITY COMPARISON
    # Create overlay image and segreded masks based on the comparison.
    # Uses the "Ranked image", i.e. the histogramm was normalized by replacing 
    # the pixel intensity by its rank in the image intensity profile.
    rkA, rkB = rank_pixels(imageA, boolmsk), rank_pixels(imageB, mask=boolmsk)
    im_dict['RkA'] = rkA = np.round(utils.image_positivehist(rkA, normalize='8bit')[0],decimals=0).astype(np.float16)
    im_dict['RkB'] = rkB = np.round(utils.image_positivehist(rkB, normalize='8bit')[0],decimals=0).astype(np.float16)

    # >> CORRELATION OF RATIO
    # > Calculation of correlation.
    pearson, spearman = images_calculate_colocalisation(imageA, imageB, mask=msksegm)
    # > Adding to the dictionary
    corr_dict = {'PCorr': pearson["PCC"], 
                    'PCerr': pearson["PCI"], 
                    'R^2': pearson["PCCS"], 
                    'SCC': spearman["SCC"], 
                    'SCI': spearman["SCI"], 
                    'SCCS': spearman["SCCS"], 
                    'SCCbeta0': spearman["SCCbeta0"], 
                    'SCCbeta1': spearman["SCCbeta1"]}
    corr_labels=['Pearson correlation', 'PCC Error', 'PCC R squarred','Spearman correlation coefficient', 'SCC C.I.', 'SCC R squarred','SCC Beta0','SCC Beta1']

    # > Flatten equalized image
    mskrkflatA, mskrkflatB = rkA[boolmsk], rkB[boolmsk]
    
    # > Inline plot of the scattergram
    _ = images_scattergram(imageA, imageB, plot = test)
    regression = images_scattergram(mskrkflatA, mskrkflatB, plot = test)
    
    # > Create dictionary to store the statistics
    export_corr = {
        "reg_x": regression[0],
        "reg_y": regression[1],
        "reg_yest": regression[2],
        "reg_yerr": np.abs(regression[3]-regression[2])
    }
    
    # > Create dictionary to store the metrics
    # > List with abbreviation and labels.
    value_dict = {**hist_dict, **corr_dict}
    value_labels = hist_labels + corr_labels
    abbr_to_label_val = {abbr: label for abbr, label in zip(value_dict.keys(), value_labels)}
    
    return im_dict, value_dict, abbr_to_label_val, export_corr

# TODO This name does not make sense.
def acceptable_values(Z,frac=0.1):
    """
    Determine acceptable values in a matrix.
  
    This function determines the distribution of y values taken in a matrix "Z" and returns the y which are at a 
    percentage distance "frac" of the distribution of value. In order to eliminate infinite value which can appear, 
    all the values 100 times superior to the median are deleted.
  
    Parameters:
    Z (np.ndarray): The input matrix.
    frac (float): The fraction of the distribution to consider. Default is 0.1.
  
    Returns:
    Tuple[np.ndarray, float, float, float]: A tuple containing the input matrix, the minimum y value, the maximum y value, 
    and the median of the y values.
    """

    n=np.size(Z,0)
    m=np.size(Z,1)
    y1=np.zeros((n))
    median_1=0
    median_2=0
    for i in range(0,n):
        y1_i_sorted=np.sort(Z[:,i])
        y1+=y1_i_sorted
        median_1+=statistics.median(y1_i_sorted)

            
    y2=np.zeros((m))
    for i in range(1,m):
        y2_i_sorted=np.sort(Z[:,i])
        y2+=y2_i_sorted
        median_2+=statistics.median(y2_i_sorted)
        
    min_y1=y1[int(frac*n)]/n
    max_y1=y1[int((1-frac)*n)]/m
    
    min_y2=y2[int(frac*n)]/n
    max_y2=y2[int((1-frac)*n)]/m
    
    return (Z,(min_y1+min_y2)/2,(max_y1+max_y2),(median_1/n+median_2/m)/2)


#TODO merge with the one for Estack into utils.
def dead_pixel(y: List[float], range_y: List[float], median: float) -> List[float]:
    """
    Replace any values outside a given range with a median value.

    This function checks each value in a given list. If a value is outside a given range, 
    it replaces the value with a given median value.

    Parameters:
    y (List[float]): The list of values to check.
    range_y (List[float]): The range to check within.
    median (float): The value to replace any out-of-range values with.

    Returns:
    List[float]: The list of values, with any out-of-range values replaced by the median value.
    """
    n=len(y)
    for i in range(0,n) :
        if not utils.is_included(y[i],range_y):
            y[i]=median
        
    return y

# TODO this name does not make sense.
def image_threshold_hue(image: np.ndarray, low_hue: int, high_hue: int) -> np.ndarray:
    """
    Threshold an RGB image based on the hue value.
    The hue value should be given as in imageJ color threshold, i.e. as a value between 1 and 255.

    Args:
        image (np.ndarray): Input RGB image (n,m,3).
        low_hue (int): Lower bound for the hue range.
        high_hue (int): Upper bound for the hue range.

    Returns:
        np.ndarray: The thresholded image (n,m,3).
        np.ndarray: The thresholded image (n,m).
    """
    # Convert the image to HSV
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Define the lower and upper bounds for the hue
    lower_bound = np.array([int(low_hue/255*180), 125, 125])
    upper_bound = np.array([int(high_hue/255*180), 255, 255])

    # Threshold the HSV image to get only the colors in the hue range
    mask = cv2.inRange(hsv, lower_bound, upper_bound)

    # Bitwise-AND the mask and the original image
    result = cv2.bitwise_and(image, image, mask=mask)

    return result, mask

def test_image_threshold_hue():
    # Create a dummy RGB image
    image = np.zeros((100, 100, 3), dtype=np.uint8)

    # Call the function with the dummy image and a hue range of 0-180
    result = image_threshold_hue(image, 0, 180)

    # Check that the result is a numpy array of the same shape as the input
    assert isinstance(result, np.ndarray)
    assert result.shape == image.shape

    # Check that the result is a valid image (values are in range 0-255)
    assert (result >= 0).all() and (result <= 255).all()
    
    # Load the image as BGR
    moduledir = os.getcwd()
    image_path=utils.path_join(moduledir,'Test/Image/trshhue/OverlayAB_MnrNir_Mnred_Nired_NCM.tif',dt='f')
    BGRimage = cv2.imread(image_path)
    RGBimage = cv2.cvtColor(BGRimage, cv2.COLOR_BGR2RGB)
    
    # Apply the threshold_hue function
    # R: 0-25 Y: 25-50 G: 60-90
    result, mask = image_threshold_hue(BGRimage, 25, 60)
    
    # Convert BGR result to RGB for plotting
    RGBmasked = cv2.cvtColor(result, cv2.COLOR_BGR2RGB)

    # Plot the original and thresholded images
    plt.figure(figsize=(10, 5))
    
    plt.subplot(1, 2, 1)
    plt.imshow(RGBimage)
    plt.title('Original Image')
    
    plt.subplot(1, 2, 2)
    plt.imshow(RGBmasked)
    plt.title('Thresholded Image')
    
    plt.figure(figsize=(10, 10))
    plt.imshow(cv2.cvtColor(result, cv2.COLOR_RGB2GRAY), cmap='gray')
    plt.title('As grayscale')
    
    # Save the result
    output_path=utils.path_join(moduledir,'Test/Image/trshhue/result.tif', dt='f')
    cv2.imwrite(output_path, RGBmasked)
    
def images_scattergram(flatA: np.ndarray, flatB: np.ndarray, plot: bool = False, mask : np.ndarray = None) -> Union[pd.DataFrame, Tuple[float, List[np.ndarray]]]:
    """
    Perform a regression analysis on two flattened and centered images, optionally plot the regression and calculate residuals.

    Args:
        flatA (np.ndarray): The flattened and centered version of the first image.
        flatB (np.ndarray): The flattened and centered version of the second image.
        plot (bool, optional): If True, a regression plot is created. Defaults to False.

    Returns:
        Union[pd.DataFrame, Tuple[float, List[np.ndarray]]]: Returns a DataFrame containing the outliers if residuals is True. Otherwise, returns a tuple containing the R-squared value and a list of arrays for the regression plot.
    """
    
    if not isinstance(flatA,np.ndarray) or not isinstance(flatB,np.ndarray) :
        return [np.nan] * 5 
    
    if not mask :
        [flatA,flatB],mask=utils.image_well_defined([flatA,flatB],axis=0)
    
    msk = mask==1
    mskflatA,mskflatB = flatA[msk], flatB[msk]
    
    def preprocess(im1,im2):
        med1, med2  = np.median(im1), np.median(im2)
        std1, std2 = np.std(im1), np.std(im2)
        nooutliers = (im1 > 0) & (im1 < (med1 + 3*std1)) & (im2> 0) & (im2 < (med2 + 3*std2))        
        im1=im1[nooutliers].copy()
        im2=im2[nooutliers].copy()
        return np.mean(im1), np.std(im1), im1, np.mean(im2), np.std(im2), im2
    
    meanA, stdA, mskflatA, meanB, stdB, mskflatB = preprocess(mskflatA,mskflatB)

    # Add column filled with 1 (intercept term) to sampleA
    mskflatA_const = sm.add_constant(mskflatA)

    # Fit the OLS model
    model = sm.OLS(mskflatB, mskflatA_const)
    results = model.fit()
    coeffs = results.params
    conf_int_low = results.conf_int(alpha=0.05, cols=None)[:,0]
    conf_int_high = results.conf_int(alpha=0.05, cols=None)[:,1]

    # Get the R-squared value
    r_squared = results.rsquared
    
    # Select a random X% of the pixels for the export
    X=1
    num_pixels = mskflatA.size
    num_sample = int(num_pixels*X)
    # for the case where there is less and 1000 pixels
    if len(str(num_sample))<3:
        num_sample = mskflatA.size
    idx = np.random.choice(num_pixels, num_sample, replace=False)
    sampleA, sampleB = quick_sort(mskflatA[idx],mskflatB[idx])

    # Create the regression line
    x = sampleA
    y_est = coeffs[1] * x + coeffs[0]

    # Calculate the 95% confidence interval
    y_lower = conf_int_low[1] * x + conf_int_low[0]
    y_upper = conf_int_high[1] * x + conf_int_high[0]

    if plot:
        plt.subplots()
        # Create the scatter plot
        x_m=np.mean(mskflatA)
        x_s=np.sqrt(np.mean((mskflatA - x_m) ** 2))
        y_m=np.mean(mskflatB)
        y_s=np.sqrt(np.mean((mskflatB - y_m) ** 2))
        plt.hexbin((flatA-x_m)/x_s, (flatB-y_m)/y_s, gridsize=50, cmap='inferno', bins='log')
        plt.colorbar(label='log10(N)')

        # plt.scatter((flatA-x_m)/x_s, (flatB-y_m)/y_s, s=20)
        plt.plot((x-x_m)/x_s, (y_est-y_m)/y_s, color='red', linewidth=1)

        # Plot the confidence interval
        plt.fill_between((x-x_m)/x_s, (y_lower-y_m)/y_s, (y_upper-y_m)/y_s, color='gray', alpha=0.5)
        plt.title(f'R-squared: {r_squared:.2f}')
        plt.show()

    return [sampleA, sampleB, y_est, y_lower, y_upper]

# This is a wrapper.
def quick_sort(*arrays: np.ndarray) -> List[np.ndarray]:
    """
    Sort an arbitrary number of arrays together by the values in the first array.

    Parameters:
    *arrays (np.ndarray): The arrays to sort.

    Returns:
    List[np.ndarray]: A list containing the sorted arrays.
    """
    # Combine arrays into list of tuples
    points = list(zip(*arrays))
    
    # Sort list of tuples by first array
    points = sorted(points, key=lambda point: point[0])
    
    # Split list of tuples into separate arrays
    sorted_arrays = [np.array(x) for x in zip(*points)]
    return sorted_arrays


def XPEEM_2i_exportCorrOrigin(folder_Origin, arg_SourceIm, name_ROI, name_Short, name_Chem, corr):
    """
    This function processes data from two different file structures.
    It is a wrapper for the sheet creation of the correlation plots in origin (Correlation, scattergramm)
    
    """
    # Extract the necessary values from the variables dictionary
    reg_x,reg_y = corr["reg_x"], corr["reg_y"]

    # Extract correlation vs mask to origin
    loc=os.getcwd()
    
    # Extract regression vs intensity to origin
    # > Get the full details of the correlation for each sample
    # > In PEEM_2i_sample_details/UncyCored2i/NCMRegr
    oplt.AddSheetOrigin(loc,'OriginPlots.opju',reg_x, reg_y,[name_Chem]*2,foldername=f'{folder_Origin}/Coloc/details/{name_Chem}',bookname=arg_SourceIm+name_Short + '--' + name_Chem,ShNam=name_ROI+name_Chem+'Regr')        

def image_mask_save(path,img,mask,trsh=2,comp=False):
        # Apply threshold and median filter of 2pt
        threshold = np.percentile(img[mask], 100-trsh)
        mask_2_percent = np.where(img > threshold, 255, 0)

        # Save mask as 8-bit image
        if not comp :
            pathmask = path.replace("grayscale_cc",f"mask/mask_{trsh}pt")
        else: 
            pathmask = path.replace("Overlay",f"mask/mask_{trsh}pt")
        io.imsave(pathmask, img_as_ubyte(mask_2_percent))
        
        return mask_2_percent==255




def images_vary_treshold(raw_A: np.ndarray, raw_B: np.ndarray, mask: np.ndarray) -> Tuple[list[float], np.ndarray]:
    """
    Calculate the Pearson and Spearman correlation for each threshold of the input images.

    Parameters:
    raw_A (np.ndarray): The first input image.
    raw_B (np.ndarray): The second input image.
    The mask of the image

    Returns:
    Tuple[List[float], np.ndarray]: A tuple containing a list of Pearson and Spearman correlation values for each threshold and the normalized version of the first input image.
    """
    # Assert input
    assert isinstance(mask,np.ndarray)
    assert raw_A.shape == raw_B.shape
    assert raw_A.shape == mask.shape
    
    # Initialize an empty list to store the correlation values
    avg_normA =utils.image_positivehist(raw_A, normalize='8bit', mask=mask)
    avg_normB =utils.image_positivehist(raw_B, normalize='8bit', mask=mask)
    
    n=100 # 2^8bit = 256bit 
    index = np.array(range(0,n,2))

    # Obtain the initial mask of the 0-values in the image.
    if not avg_normA.shape == mask.shape :        
        mask_init = np.expand_dims(mask,axis=0)
    else:
        mask_init = mask
    initsumNonZeros=np.count_nonzero(mask_init)

    
    ptmask=[]
    ptlist=[]
    sumlist = []
    PCC_list = []
    PCI_list = []
    PCCS_list = []
    slope_list = []
    intercept_list = []
    SCC_list = []
    SCI_list = []
    SCCS_list = []
    meanICQ_list = []
    intICQ_list = []
    
    savetrsh = 0
    
    for indi in index:
        # Find the pixels intensity which correspond to x% of the mask
        pt = indi/n
        trsh = array_find_contrast(avg_normA, pt_low=pt, pt_high=0, mask = mask_init)[0]
        ptlist.append(pt)
        ptmask.append(trsh)
        
        # Measure the size of the mask at this treshold value.
        mask_threshold = np.array(avg_normA > trsh) * mask_init
        summask = np.sum(mask_threshold)
        
        # Register the relative size of the mask
        sumlist.append(summask)
        
        # Requires at least 10 pixels for reliable comparison.
        trshpxlcount = 10
        if summask > trshpxlcount and trsh >= savetrsh :
            savetrsh = trsh
            # Calculate the colocalisation for the masked images
            pearson, spearman, _, ICQ = images_calculate_colocalisation(avg_normA, avg_normB, mask = mask_threshold)
            PCC_list.append(pearson["PCC"])
            PCI_list.append(pearson["PCI"])
            PCCS_list.append(pearson["PCCS"])
            intercept_list.append(pearson["PCCbeta0"])
            slope_list.append(pearson["PCCbeta1"])
            SCC_list.append(spearman["SCC"])
            SCI_list.append(spearman["SCI"])
            SCCS_list.append(spearman["SCCS"])
            meanICQ_list.append(ICQ["meanICQ"])
            intICQ_list.append(ICQ["intICQ"])
    
        else:
            PCC_list.append(np.nan)
            PCI_list.append(np.nan)
            PCCS_list.append(np.nan)
            intercept_list.append(np.nan)
            slope_list.append(np.nan)
            SCC_list.append(np.nan)
            SCI_list.append(np.nan)
            SCCS_list.append(np.nan)
            meanICQ_list.append(np.nan)
            intICQ_list.append(np.nan)
    
    pearson_vals = {"PCC": np.array(PCC_list), "PCI": np.array(PCI_list), "PCCS": np.array(PCCS_list), "PCCbeta0": np.array(intercept_list), "PCCbeta1": np.array(slope_list)}
    spearman_vals = {"SCC": np.array(SCC_list), "SCI": np.array(SCI_list), "SCCS": np.array(SCCS_list)}
    ICQ_vals = {"meanICQ": np.array(meanICQ_list), "intICQ": np.array(intICQ_list)}
    
    return pearson_vals, spearman_vals, ICQ_vals, (initsumNonZeros - np.array(sumlist))/initsumNonZeros, avg_normA

def test_images_vary_treshold():
    # Create two pairs of Gaussian images
    imageA = gaussian_filter(np.random.rand(100, 100), sigma=3)
    imageB = gaussian_filter(imageA + np.random.rand(100, 100) * 0.1, sigma=3)
    mask = np.ones((100, 100))

    # Call images_vary_treshold
    pearson_vals, spearman_vals, ICQ_vals, index, avg_normA = images_vary_treshold(imageA, imageB, mask)

    # Check that the output lists are the correct length
    assert len(pearson_vals["PCC"]) == 50
    assert len(spearman_vals["SCC"]) == 50
    assert len(ICQ_vals["meanICQ"]) == 50

    # Check that the output arrays are the correct shape
    assert index.shape == (50,)
    assert avg_normA[0].shape == (100, 100)



    # Check that the Pearson and Spearman values are between -1 and 1, or nan.
    assert (np.array([-1 <= pearson_vals["PCC"]]) * np.array([pearson_vals["PCC"] <= 1]) + np.isnan(pearson_vals["PCC"])).all()
    assert (np.array([-1 <= spearman_vals["SCC"]]) * np.array([spearman_vals["SCC"] <= 1]) + np.isnan(spearman_vals["SCC"])).all()
    assert (np.array([-0.5 <= ICQ_vals["meanICQ"]]) * np.array([ICQ_vals["meanICQ"] <= 0.5]) + np.isnan(ICQ_vals["meanICQ"])).all()

def images_flatten(imageA: np.ndarray, imageB: np.ndarray, mask: np.ndarray=None) -> Tuple[dict, dict, np.ndarray, np.ndarray, np.ndarray, Tuple[int, int]]:
    """
    Flattens two images and applies a mask to them.
    
    If no mask is provided, the function treats zero values in the images as the mask.
    
    Parameters:
    imageA (np.ndarray): The first image to be flattened.
    imageB (np.ndarray): The second image to be flattened.
    mask (np.ndarray, optional): The mask to be applied to the images. If not provided, zero values in the images are treated as the mask.
    
    Returns:
    mskimA (np.ndarray): The flattened and masked version of imageA.
    mskimB (np.ndarray): The flattened and masked version of imageB.
    mapping (np.ndarray): A mapping of the initial position of the pixels of the image.
    shapeA (tuple): The shape of imageA.
    
    Raises:
    ValueError: If the mask is not the same size as imageA.
    AssertionError: If the shape of imageA does not match the shape of imageB or mask, or if there are zero values in the masked images.
    """

    # If no mask is provided, takes zero values as the mask.
    if not isinstance(mask, np.ndarray) :
        [imageA, imageB], mask=utils.image_well_defined(imageA,imageB)
    
    # Take the shape and check it.
    shapeA = imageA.shape
    assert shapeA == imageB.shape
    if shapeA != mask.shape:
        if shapeA[0]==1 and mask.shape[0]>1:
            mask = np.expand_dims(mask, 0)
            assert shapeA == mask.shape
        else:
            raise ValueError('Please provide a mask of the same size as imageA')
        
    # Apply the mask
    mskimA=imageA[mask==1]
    mskimB=imageB[mask==1]
       
    # >> Save a mapping of the initial position of the pixels of Im
    mapping = mask
    
    return mskimA, mskimB, mapping, shapeA


def images_calculate_colocalisation(imageA: np.ndarray, imageB: np.ndarray, mask: np.ndarray=None) -> Tuple[dict, dict, np.ndarray, np.ndarray, np.ndarray, Tuple[int, int]]:
    """
    Calculate the Pearson and Spearman correlation between two images and return the correlation, Manders' coefficients, Intensity Correlation Quotient, flattened and centered versions of imageA and imageB, mapping of the initial position of the pixels, and the shape of the images.

    Args:
        imageA (np.ndarray): The first image.
        imageB (np.ndarray): The second image.
        mask (np.ndarray, optional): If provided, only the pixels where mask is True are considered. If not provided, all pixels are considered.

    Returns:
        Tuple[dict, dict, dict, dict]: Returns a tuple containing the Pearson correlation, Spearman correlation, Manders' coefficients, Intensity Correlation Quotient
    """     
    # If no mask is provided, takes non zero values as the image.
    [imageA, imageB], mask=utils.image_well_defined([imageA,imageB],axis=0)
    
    # Check that the mask is not 1 for less than 5 pixels. 
    if  np.count_nonzero(mask) < 5:
        return np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
    
    mskimA, mskimB,_,_ = images_flatten(imageA, imageB, mask)
    
    # >> Calculate Pearson correlation, error bars and square
    # Measures the correlation between the intensities of two channels. 
    # A PCC of 1 means perfect positive correlation, -1 means perfect negative correlation, and 0 means no correlation.
    # Note: interval of confidence based on Fischer transformation
    # Ref. C. J. Kowalski, "On the Effects of Non-Normality on the Distribution of the Sample Product-Moment Correlation Coefficient" Journal of the Royal Statistical Society. Series C (Applied Statistics), Vol. 21, No. 1 (1972), pp. 1-12.
    # >> Calculate Pearson correlation 
    # > Data normalisation
    def preprocess(im1,im2):
        med1, std1 = np.median(im1), np.std(im1)
        med2, std2 = np.median(im2), np.std(im2)
        nooutliers = (im1 > 0) & (im1 < (med1 + 3*std1)) & (im2> 0) & (im2 < (med2 + 3*std2))        
        im1=im1[nooutliers]
        im2=im2[nooutliers]
        return np.mean(im1), np.std(im1), im1, np.mean(im2), np.std(im2), im2

    meanA, stdA, mskimA, meanB, stdB, mskimB = preprocess(mskimA,mskimB)
    normA = (mskimA-meanA)/(stdA)
    normB = (mskimB-meanB)/(stdB)
    
    # > Regression    
    PCCbeta1, PCCbeta0, PCC, _, std_err = linregress(normA, normB)
    # > Fisher transformation
    fisher_r = np.arctanh(PCC)
    # > Calculate the standard error
    std_err = 1 / np.sqrt(len(mskimA) - 3)
    # > Calculate the confidence interval
    z = stats.norm.ppf(1 - (1 - 0.99) / 2)
    PCI = np.tanh(fisher_r + z * std_err) - PCC
    # Store in a dictionary
    pearson = {
        "PCC": PCC,
        "PCI": PCI,
        "PCCS": PCC**2,
        "PCCbeta0": PCCbeta0,
        "PCCbeta1": PCCbeta1
        }
    
    # >> Calculate Spearman's correlation coefficient. R
    # Ranges from -1 to 1. 
    # 1 means a perfect positive monotonic relationship
    # -1 means a perfect negative monotonic relationship
    # 0 means no relationship.
    # It is more robust than Pearson to outliers, but it does not prove linearity.
    rkdimageA = utils.image_positivehist(image_rankpixels(imageA), normalize='8bit')+1
    rkdimageB = utils.image_positivehist(image_rankpixels(imageB), normalize='8bit')+1
    if rkdimageA.shape != mask.shape:
        rkdimageA=rkdimageA[0]
        rkdimageB=rkdimageB[0]
    rkdmskimA=rkdimageA[mask==1]
    rkdmskimB=rkdimageB[mask==1]

    SCCbeta1, SCCbeta0, SCC, _, SCC_std_err = linregress(rkdmskimA, rkdmskimB)
    # Fisher transformation
    fisher_r = np.arctanh(SCC)
    
    # Calculate the standard error
    std_err = 1 / np.sqrt(len(rkdmskimA) - 3)
    
    # Calculate the confidence interval with the Fisher transformation
    z = stats.norm.ppf(1 - (1 - 0.99) / 2)
    SCI = np.tanh(fisher_r + z * std_err) - SCC

    # Store in a dictionary
    spearman = {
        "SCC": SCC,
        "SCI": SCI,
        "SCCS": SCC**2,
        "SCCbeta0": SCCbeta0,
        "SCCbeta1": SCCbeta1,
        }
    
    return pearson, spearman
    
def test_images_calculate_colocalisation():
    # Create two pairs of Gaussian images
    imageA_high = gaussian_filter(np.random.rand(100, 100), sigma=3)
    imageB_high = gaussian_filter(imageA_high + np.random.rand(100, 100) * 0.1, sigma=3)

    imageA_low = gaussian_filter(np.random.rand(100, 100), sigma=3)
    imageB_low = gaussian_filter(np.random.rand(100, 100), sigma=3)

    # Apply calculate_colocalisation
    pearson_high, spearman_high = images_calculate_colocalisation(imageA_high, imageB_high)
    pearson_low, spearman_low = images_calculate_colocalisation(imageA_low, imageB_low)

    # Create a comparison table
    data_high = {
        "Pearson Correlation": [pearson_high["PCC"]],
        "Pearson Confidence Interval": [pearson_high["PCI"]],
        "Pearson Correlation Square": [pearson_high["PCCS"]],
        "Spearman Correlation": [spearman_high["SCC"]],
        "Spearman Confidence Interval": [spearman_high["SCI"]],
        "Spearman Correlation Square": [spearman_high["SCCS"]],
    }
    df_high = pd.DataFrame(data_high, index=['High Correlation'])

    data_low = {
        "Pearson Correlation": [pearson_low["PCC"]],
        "Pearson Confidence Interval": [pearson_low["PCI"]],
        "Pearson Correlation Square": [pearson_low["PCCS"]],
        "Spearman Correlation": [spearman_low["SCC"]],
        "Spearman Confidence Interval": [spearman_low["SCI"]],
        "Spearman Correlation Square": [spearman_low["SCCS"]],
    }
    df_low = pd.DataFrame(data_low, index=['Low Correlation'])

    df = pd.concat([df_high, df_low])
    print(df.T)
    
    # Plot a comparison before/after
    fig, axs = plt.subplots(2, 2, figsize=(10, 10))
    axs[0, 0].imshow(imageA_high, cmap='gray')
    axs[0, 0].set_title('Image A High')
    axs[0, 1].imshow(imageB_high, cmap='gray')
    axs[0, 1].set_title('Image B High')
    axs[1, 0].imshow(imageA_low, cmap='gray')
    axs[1, 0].set_title('Image A Low')
    axs[1, 1].imshow(imageB_low, cmap='gray')
    axs[1, 1].set_title('Image B Low')
    plt.show()
        
def image_rankpixels(image: np.ndarray, mask = None) -> np.ndarray:
    """
    Replace each pixel's intensity with its rank.

    Parameters:
    image (np.ndarray): The input image.

    Returns:
    np.ndarray: The ranked image.
    """
    if not isinstance(mask,np.ndarray) :
        [image], mask = utils.image_well_defined(image)
        
    # Flatten the image into a 1D array and get the sorted indices
    # Only consider pixels where the mask is True
    ranks = rankdata(image[mask==1])-1

    # Create an empty array of the same shape as the image
    ranked_image = np.zeros_like(image)

    # Fill the ranked_image array with ranks
    # Only fill at locations where the mask is True
    np.place(ranked_image, mask==1, ranks.astype(image.dtype))

    return ranked_image

def test_image_rankpixels():
    mean1 = (40, 40)
    size=100
    sigma=10
    num_images=1
    # Create empty image stacks
    image = np.zeros((num_images, size, size))

    # Create a grid of (x, y) coordinates
    x, y = np.meshgrid(np.linspace(0, size-1, size), np.linspace(0, size-1, size))

    # Create a 2D Gaussian distribution
    d = np.sqrt((x-mean1[0])**2 + (y-mean1[1])**2)
    image = np.exp(-(d**2 / (2.0 * sigma**2)))
    

    # Apply rank_image
    ranked_image = image_rankpixels(image)

    # Check that the ranked image is not empty
    assert ranked_image.size != 0, "The ranked image is empty."

    # Check that the ranked image has the same dimensions as the input image
    assert ranked_image.shape == image.shape, "The ranked image has different dimensions than the input image."

    # Plot a comparison before/after
    fig, axs = plt.subplots(2, 2, figsize=(10, 10))
    axs[0, 0].imshow(image, cmap='gray')
    axs[0, 0].set_title('Original Image')
    axs[0, 1].imshow(ranked_image, cmap='gray')
    axs[0, 1].set_title('Ranked Image')
    
    # Plot the intensity values
    axs[1, 0].hist(image.flatten(), bins=50, color='blue', alpha=0.7)
    axs[1, 0].set_title('Original Image Intensity Values')
    axs[1, 1].hist(ranked_image.flatten(), bins=50, color='orange', alpha=0.7)
    axs[1, 1].set_title('Ranked Image Intensity Values')
    
    plt.tight_layout()
    plt.show()
    
    # Create a test image
    image = np.array([[1, 2, 3], [4, -1, 6], [7, 8, 9]])
    
    # Create a mask where the middle pixel is masked out
    mask = np.array([[True, True, True], [True, False, True], [True, True, True]])
    
    # Apply the function
    ranked_image = image_rankpixels(image, mask)
    
    # Check that the masked pixel was not taken into account
    assert ranked_image[1, 1] == 0
    
    # Check that the other pixels were ranked correctly
    assert np.all(ranked_image[mask] == np.arange(8))




def test_images_scattergram():
    # Create two pairs of Gaussian images
    imageA = gaussian_filter(np.random.rand(100, 100), sigma=3)
    imageB = gaussian_filter(imageA + np.random.rand(100, 100) * 0.1, sigma=3)
    flatA = imageA.flatten()
    flatB = imageB.flatten()

    # Call images_scattergram
    result = images_scattergram(flatA, flatB, plot=True)

    # Check that the output is a list of the correct length
    assert isinstance(result, list)
    assert len(result) == 5

    # Check that each element of the output is a numpy array
    for array in result:
        assert isinstance(array, np.ndarray)

def update_excel_sheet(df : pd.DataFrame(), excel_path : str, sheet : str):
    """
    Updates the "sheet" sheet in an existing Excel file with a DataFrame.

    Args:
        df (DataFrame): The DataFrame to write to the Excel file.
        excel_path (str): The path to the Excel file.

    Returns:
        None
    """
    if df.empty:
        raise ValueError("The dataframe argument should contain values")

    # Load the existing Excel file
    book = load_workbook(excel_path)

    # Check if the sheet exists and remove it
    if sheet in book.sheetnames:
        del book[sheet]
        book.save(excel_path)

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        # Write the DataFrame to the "summary" sheet
        df.to_excel(writer, sheet_name=sheet, index=False)

    print(f'Successfully updated the "summary" sheet in {excel_path}')

def test_update_excel_sheet():
    # Create a DataFrame
    df = pd.DataFrame({
        'A': [1, 2, 3],
        'B': [4, 5, 6],
        'C': [7, 8, 9]
    })

    # Define the sheet names
    sheet1 = 'summary'
    sheet2 = 'extra'

    # Create a temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp:
        excel_path = temp.name

        # Write the DataFrame to the Excel file in two sheets
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet1, index=False)
            df.to_excel(writer, sheet_name=sheet2, index=False)

        # Call the function
        update_excel_sheet(df, excel_path, sheet1)

        # Check if the file exists
        assert os.path.exists(excel_path), 'Excel file does not exist'

        # Load the Excel file
        df_loaded = pd.read_excel(excel_path, sheet_name=sheet1)

        # Check if the loaded DataFrame is equal to the original DataFrame
        pd.testing.assert_frame_equal(df, df_loaded)

        print('Test passed')

def array_find_contrast(im, pt_low=0.1, pt_high=0.1, cnt_low= None, cnt_high=None, mask: np.ndarray = None):
    """
    Find the contrast of an image array.

    Parameters
    ----------
    im : ndarray
        The input image array (single image).
    pt_low : float, optional
        The lower percentile for contrast adjustment. Defaults to 0.1.
    pt_high : float, optional
        The upper percentile for contrast adjustment. Defaults to 0.1.
    cnt_low : float, optional
        The lower contrast limit. If provided, overrides pt_low. Defaults to None.
    cnt_high : float, optional
        The upper contrast limit. If provided, overrides pt_high. Defaults to None.

    Returns
    -------
    min_ : float
        The minimum pixel value for contrast adjustment.
    max_ : float
        The maximum pixel value for contrast adjustment.
    """
    
    # Assert the inputs arrays
    assert isinstance(im,np.ndarray)
    imshape = im.shape
    ndim = len(imshape) 
    if ndim == 3:
        assert imshape[0] == 1
    elif ndim == 2:
        np.expand_dims(im, 0)
    else:
        raise ValueError('array_find_contrast Please provide an image to the function ')
    
    if np.all(np.isnan(im)) or np.all(np.isinf(im)) :
        return np.nan, np.nan
    
    # Calculate the mask if it is not provided.
    if not isinstance(mask,np.ndarray) : 
        [im], mask = utils.image_well_defined(im)
    else:
        assert mask.shape == im.shape
        
        
    fmsk=mask.flatten()
    fim=im.flatten()
    fimsk=fim[fmsk == 1]
    
    if not (fimsk!=0).all() :
        print(fimsk == 0)
        raise ValueError('The mask is not good')

    # Compute the histogram of the non-zero values
    non_zero_min, non_zero_max = np.min(fimsk), np.max(fimsk)
    histogram = np.histogram(im, bins=256, range=(non_zero_min, non_zero_max))[0]
    bin_size = (non_zero_max - non_zero_min) / 256

    # Compute the threshold for determining the output min and max bins
    pixel_count = np.count_nonzero(im)
    max_cnt = pixel_count - 10

    limitlow = cnt_low if cnt_low else pixel_count*pt_low
    limithigh = cnt_high if cnt_high else pixel_count*pt_high
        
    # Determine the output min and max bins
    cumulative_counts = np.cumsum(histogram)
    hmin = next(i for i, ccount in enumerate(cumulative_counts) if ccount > limitlow or ccount > max_cnt)
    cumulative_counts_rev = np.cumsum(histogram[::-1])
    hmax = 255 - [i for i, ccount in enumerate(cumulative_counts_rev) if ccount > limithigh or ccount > max_cnt][0] 

    # Compute the output min and max pixel values
    min_, max_ = (non_zero_min + hmin * bin_size, non_zero_min + hmax * bin_size) if hmax >= hmin else (non_zero_min, non_zero_max)
    min_, max_ = (min_, max_) if min_ != max_ else (non_zero_min, non_zero_max)

    return min_, max_

def test_array_find_contrast_count_based():
    # Create a test image
    im = np.random.rand(100, 100)

    # Call the function with count-based limits
    min_, max_ = array_find_contrast(im, cnt_low=5, cnt_high=5)

    # Check the results
    assert min_ >= np.min(im)
    assert max_ <= np.max(im)
    assert min_ < max_

    # Call the function with percentile-based limits
    min_, max_ = array_find_contrast(im, pt_low=0.3, pt_high=0.8)

    # Check the results
    assert min_ >= np.min(im)
    assert max_ <= np.max(im)
    assert min_ < max_
    
    print('test_array_find_contrast_count_based run without issue')

def image_adjust_contrast(image_path, output_path):
    """
    Load an image, adjust its contrast based on its histogram, and save the processed image.

    Parameters
    ----------
    image_path : str
        Path to the input image.
    output_path : str
        Path to save the processed image.

    Returns
    -------
    None
    """
    # Check if the input folders exist
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"The input folder {image_path} does not exist.")
    if not os.path.exists(os.path.dirname(output_path)):
        raise FileNotFoundError(f"The input folder {output_path} does not exist.")
    

    # Load the image
    im = cv2.imread(image_path, -1)
    im_type = im.dtype
    im=np.nan_to_num(im,nan=0,posinf=0)

    # If the image is color, convert it to grayscale
    if len(im.shape) == 3 and im.shape[2] == 3:
        im = 0.3 * im[:,:,2] + 0.59 * im[:,:,1] + 0.11 * im[:,:,0]
        im = im.astype(im_type)
    
    
    # Find the contrast of the image array
    min_, max_ = array_find_contrast(im)
    if max_ - min_ == 0 :
        print(f'image_adjust_contrast : max_ and min_ are the same, the image {image_path} is a constant value.')
        
    # Adjust the contrast of the image
    imr = np.zeros_like(im)

    imr[im > 0] = (im[im > 0] - min_) / (max_ - min_) * 255
    imr[(im < min_) * (im > 0)] = 1
    imr[im > max_] = 255
    
    # Save the contrast-adjusted image
    cv2.imwrite(output_path, imr)


def test_image_adjust_contrast():
    # Create a synthetic image
    im = np.random.randint(0, 256, (100, 100), dtype=np.uint8)

    # Save the synthetic image to a temporary file
    input_fd, input_path = tempfile.mkstemp(suffix='.tif')
    cv2.imwrite(input_path, im)
    os.close(input_fd)  # Close the file descriptor

    # Run the function on the temporary file
    output_fd, output_path = tempfile.mkstemp(suffix='.tif')
    image_adjust_contrast(input_path, output_path)
    os.close(output_fd)  # Close the file descriptor

    # Check that the output file exists and is not empty
    assert os.path.exists(output_path)
    assert os.path.getsize(output_path) > 0

    time.sleep(1)

    # Clean up the temporary files
    os.remove(input_path)
    os.remove(output_path)
    
    print('test_image_adjust_contrast ran without mistake')


def rank_pixels(im: np.ndarray, mask: Optional[np.ndarray] = None) -> Image:
    """
    Ranks the pixels in an image according to their intensity.

    Parameters:
    im (np.ndarray): The input image.
    mask (np.ndarray, optional): A binary mask of the same shape as `im`. If provided, only the pixels where `mask` is True are ranked.

    Returns:
    Image: The ranked image, where each pixel's intensity is replaced by its rank.
    """
    if not isinstance(mask, np.ndarray):
        [im], binmsk = utils.image_well_defined(im)
        mask = binmsk == 1  # Convert to boolean
    else: 
        assert mask.shape == im.shape
        
    # Apply the mask to the pixels
    masked_pixels = im[mask]

    # Rank the non-zero pixels
    ranked_pixels = rankdata(masked_pixels, method='average')

    # Create a new array with the same shape as the original image
    ranked_image_pixels = np.zeros_like(im)

    # Assign the ranks to the corresponding positions in the new array
    ranked_image_pixels[mask] = ranked_pixels

    return ranked_image_pixels

def test_rank_pixels():
    # Create a test image and mask
    image = np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
    mask = np.array([[1, 0, 1], [0, 1, 0], [1, 0, 1]], dtype=bool)

    # Call the rank_pixels function
    ranked_image = rank_pixels(image, mask)

    # Check that the output is an instance of Image
    assert isinstance(ranked_image, np.ndarray)

    # Check that the pixel values in the ranked image are correct
    expected_ranked_image = np.array([[1, 0, 2], [0, 3, 0], [4, 0, 5]], dtype=np.uint8)
    np.testing.assert_array_equal(np.array(ranked_image), expected_ranked_image)

    # Test with no mask provided
    ranked_image_no_mask = rank_pixels(image)
    assert isinstance(ranked_image_no_mask, np.ndarray)

def run_imagej_macro(macro_path: str, input_args: List[str], holdtime:int = 0) -> None:
    """
    Run an ImageJ macro with given arguments.

    Parameters:
    macro_path (str): The path to the ImageJ macro to run.
    input_args (List[str]): A list of arguments to pass to the ImageJ macro.

    Returns:
    None
    """
    init_dir=os.getcwd()
    imagej_path = "C:/Users/lelotte_b/Desktop/ImageJ/ImageJ.exe" 
    os.chdir(os.path.dirname(imagej_path))
    java_executable = "java"  # Ensure 'java' is in your PATH, or provide the full path to the java executable
    start_jar = "-Xmx512m -jar ij.jar"
    args = ",".join(input_args)
    command = [
        java_executable,
        *start_jar.split(),
        "--headless",
        "-macro",
        macro_path,
        args
    ]
    success = True
    _=subprocess.run(command)
    
    if 'distorsion_correction' in macro_path :
        # Wait until the distorded images are processed.
        # If the programm stops for wathever reason, exit the loop and returns False
        nrIm = holdtime
        dstrd_folder = input_args[2]
        still_time = 0
        nrImFolder = 0
        while(nrImFolder<nrIm and still_time < nrIm):
            time.sleep(nrIm/10)
            
            # Get the number of image in the folder
            nrImFolderi = len(os.listdir(dstrd_folder))
            
            # still_time is used to detect if imageJ has stopped working.
            if nrImFolder == nrImFolderi :
                still_time += 0.1/nrIm
            nrImFolder = nrImFolderi
        if not still_time < nrIm :
            success = False
        time.sleep(2)
        
    elif holdtime > 0 :
        # Wait for holdtime seconds
        time.sleep(holdtime*2)
    
    # Define the command to close ImageJ
    command_close_imagej = ['taskkill', '/IM', 'java.exe', '/F']
    
    # Close ImageJ
    try:
        subprocess.run(command_close_imagej, check=True)
        print("ImageJ closed successfully.")
    except subprocess.CalledProcessError as e:
        print(f"Error closing ImageJ: {e}")
    
    os.chdir(init_dir)
    
    if success :
        print(f"The IJ macro has finished ({macro_path}), closing imageJ.")
        return True
    else : 
        print("The IJ macro has encountered an error.")
        return False


def IJ_col_and_scalebar(inputfolder:str, outputfolder:str ) -> None:
    """
    Run an ImageJ plugin which adds a colormap and a scalebar with given arguments and 
    set the current directory back to the original one.

    Parameters:
    inputfolder (str): The path to the folder containing the images to process.
    outputfolder (str): The path to the folder where the processed images will be saved.

    Returns:
    None
    """

    
    input_args = [inputfolder, outputfolder]
    macro_path = "C:/Users/lelotte_b/Desktop/ImageJ/plugins/image_scalebar_jet_rgb_py.ijm"
    run_imagej_macro(macro_path, input_args)
    

def IJ_bUnwarpJ(target_image: str, sourcefolder: str, version: int = 1, args="1;1;1;1;Coarse;0.01;0.001") -> None:
    """
    Run an ImageJ plugin which aligns a set of images to a target reference image using a base transform.

    Parameters:
    target_image (str): The path to the target reference image.
    base_transform (str): The path to the base transform.
    folder (str): The path to the folder containing the images to align.

    Returns:
    None
    """
    parts = args.split(";")
    assert len(parts) == 7 or ((len(parts) == 8 and (version == 6 or version == 7 or version == 8)) or (len(parts) == 9 and version == 8))
    
    # Define the ImageJ macro version (set of parameter for alignement)
    # 1 optimized for image from the same edge and same stack.
    # 2 optimized for image from different edges.
    # 3 optimized for image from the same edge and different stacks (spectrum vs 2_energies).
    
    assert version in [1,2,3,4,5,6,7,8]
    macro_path = "C:/Users/lelotte_b/Desktop/ImageJ/plugins/distorsion_correction_py"+str(version)+".ijm"

    # Create a direct transform, indirect transform and output folder.
    if sourcefolder.endswith('/') or sourcefolder.endswith('\\'):
        # Rename
        directtrsffolder = utils.path_join(sourcefolder.rstrip('/\\') + '_drtrsf')
        indirecttrsffolder = utils.path_join(sourcefolder.rstrip('/\\') + '_indrtrsf')
        outputfolder = utils.path_join(sourcefolder.rstrip('/\\') + '_undistrdd')

        # Create the output folder
        os.makedirs(directtrsffolder, exist_ok=True)
        os.makedirs(indirecttrsffolder, exist_ok=True)
        os.makedirs(outputfolder, exist_ok=True)
        prevexecution = [f for f in os.listdir(outputfolder) if f.lower().endswith(('.tif'))]
        for f in prevexecution:
            os.remove(utils.path_join(outputfolder)+f)
    else:
        raise ValueError('The folder name must end with a slash.')
     
    input_args = [target_image, sourcefolder, outputfolder, directtrsffolder, indirecttrsffolder, args]
    
    # Get a list of all image files in the source folder
    image_files = [f for f in os.listdir(sourcefolder) if f.lower().endswith(('.tif'))]

    
    nrim = len(image_files)

    # Run the ImageJ macro with the current image file
    run_imagej_macro(macro_path, input_args, holdtime = nrim)
    
def IJ_applytrsf(inputfolder,transform,reftransform, targetpath):
    # >> Estimate execution time
    # Get a list of all image files in the source folder
    image_files = [f for f in os.listdir(inputfolder) if f.lower().endswith(('.tif'))]
    nrim = len(image_files)
    
    # >> Generate/clean output folder
    outputUnWarpJ=utils.path_join(inputfolder.rstrip('/\\') + '_undistrdd')
    outputUnWarpJNew=utils.path_join(inputfolder.rstrip('/\\') + '_undistrdd0')
    if os.path.exists(outputUnWarpJ) and not os.path.exists(outputUnWarpJNew) :
        os.rename(outputUnWarpJ, outputUnWarpJNew)
    outputfolder = outputUnWarpJ
    os.makedirs(outputfolder, exist_ok=True)
    prevexecution = [f for f in os.listdir(outputfolder) if f.lower().endswith(('.tif'))]
    for f in prevexecution:
        os.remove(utils.path_join(outputfolder)+f)
    
    # >> Smooth transforms
    input_args = [transform,reftransform, outputfolder, inputfolder, targetpath]
    macro_path = "C:/Users/lelotte_b/Desktop/ImageJ/plugins/consistency_applyrectifiedtransform.ijm"
    run_imagej_macro(macro_path, input_args, holdtime = nrim/4)
    
    # >> Apply linear alignement
    # input_args = [outputfolder]
    # macro_path = "C:/Users/lelotte_b/Desktop/ImageJ/plugins/consistency_StackReg.ijm"

    # run_imagej_macro(macro_path, input_args, holdtime = nrim/2)

def test_IJ_bUnwarpJ():
    # TEST FOLDER 0
    # Define the target image, base transform, and folder
    # Most distorded images on C K-edge.
    # BEST: v5 1-1-1 Coarse 0.001 - both masks.
    best0 = "0;1;1;1;Coarse;0.001"
    basedir=os.getcwd()
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref_C_uncy_s11009_001.tif',dt='f') # C
    folder = utils.path_join(basedir,'Test/bUnwarpJ/test_series',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 5, args=best0)

    # TEST FOLDER 1
    # Define the target image, base transform, and folder
    # C K-edge, similar to the reference.
    # BEST: v5 1-1-1 Coarse 0.001 - both masks.
    best1 = "0;1;1;1;Coarse;0.001"
    basedir=os.getcwd()
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref_C_uncy_s11009_001.tif',dt='f')
    folder = utils.path_join(basedir,'Test/bUnwarpJ/test_series1',dt='d')
    
    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 5, args=best1)   
    
    # TEST FOLDER 2
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - minimal mask for the target 10% source.
    best2 = "1;1;1;0.5;Coarse;0.001"
    basedir=os.getcwd()
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref_Co_uncy_s10997_003.tif',dt='f')
    folder = utils.path_join(basedir,'Test/bUnwarpJ/test_series2',dt='d')
    
    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 5, args=best2)
    
    # TEST FOLDER 3
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    best3 = "0;1;1;0.5;Coarse;0.001"
    basedir=os.getcwd()
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref_C_uncy_s11009_001.tif',dt='f')
    # folder = utils.path_join(basedir,'Test/bUnwarpJ/test_series3',dt='d')
    
    # Call the function
    # IJ_bUnwarpJ(target_image, folder, verion = 5, args=best3)
    
    # TEST FOLDER 4
    # Define the target image, base transform, and folder
    # The reference P_S which was problematic
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    best3 = "0;10;10;1;Coarse;0.001"
    basedir=os.getcwd()
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/raw_cy10_undistrdd/O_10cy_s10819_001.tif',dt='f') 
    folder = utils.path_join(basedir,'Test/bUnwarpJ/test_series4',dt='d')
    
    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 5, args=best4)
    
    # TEST FOLDER 5
    # Define the target image, base transform, and folder
    # The different edges for the 2 energies.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    best = "1;100;10;0.1;Very Coarse;0.001"
    basedir=os.getcwd()
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref2_C_cy10_i211020_095#001.tif',dt='f') 
    folder = utils.path_join(basedir,'Test/bUnwarpJ/test_series5',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 5, args=best)

    # UNCYCLED 1SPECTRUM ALIGNEMENT    
    # Define the target image, base transform, and folder
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref_O_uncy_i211025_034#001.tif',dt='f')
    best1 = "0;10;10;10;Coarse;0.001;1000;1" # All but C and P_S with ref_Ni_uncy_s10993_001
    best2 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # C and P_S with ref_O_uncy_s11005_001
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/raw_uncy',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 6, args=best2)
    
    # UNCYCLED 2ENERGY ALIGNEMENT
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref_O_uncy_i211025_034#001.tif',dt='f')
    best1 = "0;10;10;10;Coarse;0.001;1000;1" # All but C and P_S with ref_Ni_uncy_s10993_001
    best2 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # C and P_S with ref_O_uncy_s11005_001
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/2E_uncy',dt='d')
    
    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 6, args=best2)

    # 10 CYCLES 1SPECTRUM ALIGNEMENT    
    # Define the target image, base transform, and folder
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref_Ni_cy10_s10804_001.tif',dt='f')
    best1 = "0;10;10;10;Coarse;0.001;1000;1"  # All but C and P_S with ref_Ni_cy10_s10804_001
    best2 = "0;10;20;10;Coarse;0.001;1000;2"+","+target_image  # C and P_S with ref_O_cy10_s10819_001edge
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/raw_cy10',dt='d')
   
    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 6, args=best1)
    
    # 10 CYCLES 2ENERGY ALIGNEMENT
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/ref_Ni_cy10_s10804_001.tif',dt='f') 
    best1 = "0;10;10;10;Coarse;0.001;1000;1" # All but C and P_S with ref_Ni_cy10_s10804_001 with ve6 and 1
    best2 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # C and P_S with ref_O_cy10_s10819_001edge with ve6 and 2
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/2E_cy10',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 6, args=best1)
    
    # >> oXAS LNO OCP Raw 2ENERGY ALIGNEMENT
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/_abs_ref_OCP/test_Ni_OCP_i230818_049#001.tif',dt='f') 
    visu = "0;100;1000;1000;Coarse;0.001;250;0"+","+target_image # vizualization purpose.
    best1 = "0;1000;1000;1000;Coarse;0.001;1"+","+target_image # TMx and C with ref_Ni_OCP_s22547_055 with ve6 and 1
    best2 = "0;10;10;1000;Coarse;0.001;150;0"+","+target_image # O, S with C_uncy_s22569_117 with ve6 and 2
    best3 = "0;10;10;1000;Coarse;0.001;1000;2"+","+target_image # P_S_Cl with Co_uncy_s22549_072 with ve6 and 2
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/raw_OCP',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 7, args=best1)
    
    # >> oXAS LNO OCP 2E 2ENERGY ALIGNEMENT
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/_abs_ref_OCP/EDF_low2_edge_filtered.tif',dt='f') 
    visu = "0;10;10;10;Coarse;0.001;1000;0"+","+target_image # vizualization purpose.
    best0 = "0;10000;5000;10000;Fine;0.001;0.25;2"+","+target_image # All but Cl and Nb with O_OCP_s22568_036 with ve6 and 2
    best1 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # All but Cl and Nb with O_OCP_s22568_036 with ve6 and 2
    best2 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # Cl and Nb with P_OCP_i230818_116 with ve6 and 2
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/2E_OCP',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 7, args=best0)
    
    # >> oXAS LNO Chrg Raw 2ENERGY ALIGNEMENT
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/_abs_ref_Chrg/O_cycl_s22624_040.tif',dt='f') 
    visu = "0;10;10;10;Coarse;0.001;1000;0"+","+target_image # vizualization purpose.
    best1 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # TMx with Ni_cycl_s22613_054 with ve6 and 2
    best2 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # with O_cycl_s22624_040 with ve6 and 2
    best3 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # Nb and S with P_cycl_s22631_041 with ve6 and 2
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/raw_Chrg',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 6, args=visu)
    
    # >> oXAS LNO Chrg 2E 2ENERGY ALIGNEMENT
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/_abs_ref_Chrg/O_cycl_s22624_040.tif',dt='f') 
    visu = "0;10;10;10;Coarse;0.001;1000;0"+","+target_image # vizualization purpose.
    best1 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # Mn, Co, O and P with Ni_cycl_s22613_054 with ve6 and 2
    best2 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # Ni, Cl with O_cycl_s22624_040 with ve6 and 2
    best3 = "0;10;10;10;Coarse;0.001;1000;2"+","+target_image # Nb with P_cycl_s22631_041 with ve6 and 2
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/2E_Chrg',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 6, args=visu)

    # >> oXAS Pristine OCP 2E 2ENERGY ALIGNEMENT
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/_abs_ref_OCPPr/P_OCPPr_i230321_049.tif',dt='f') 
    visu = "0;10;10;10;Fine;0.001;1000;0"+","+target_image # vizualization purpose.
    best1 = "0;10;10;10;Fine;0.001;1000;2"+","+target_image # Co, Mn, O and P with Ni_OCP_i230321_003 with ve6 and 2 OK
    best2 = "0;100;100;100;Fine;0.001;200;2"+","+target_image # C P_OCPPr_i230321_049 with ve6 and 2
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/2E_OCPPr',dt='d')

    # Call the function
    # IJ_bUnwarpJ(target_image, folder, version = 6, args=best2)

    # >> oXAS Pristine Chrg 2E 2ENERGY ALIGNEMENT
    # Define the target image, base transform, and folder
    # The different edges.
    # BEST: v5 1-1-0.5 Coarse 0.001 - both masks.
    target_image = utils.path_join(basedir,'Test/bUnwarpJ/_abs_ref_ChrgPr/Co_ChrgPr_i230322_020.tif',dt='f') 
    visu = "0;10;10;10;Fine;0.001;1000;0"+","+target_image # vizualization purpose.
    best1 = "0;10;10;10;Fine;0.001;1000;2"+","+target_image # Ni and P? with Ni_OCP_i230321_003 with ve6 and 2 
    best2 = "0;10;10;10;Fine;0.001;1000;2"+","+target_image #  Co and O with Ni_ChrgPr_i230322_014 with ve6 and 2 
    best3 = "0;10;10;10;Fine;0.001;1000;2"+","+target_image #  C and Mn with Co_ChrgPr_i230322_020 with ve6 and 2 
    best4 = "0;100;100;100;Fine;0.001;200;0"+","+target_image #  C with Co_ChrgPr_i230322_020 with ve6 and 2 
    basedir=os.getcwd()
    folder = utils.path_join(basedir,'Test/bUnwarpJ/2E_ChrgPr',dt='d')

    # Call the function
    IJ_bUnwarpJ(target_image, folder, version = 6, args=best4)

def smoothtrsf(drtrsfpath):
    # List the files
    filelist=os.listdir(drtrsfpath)
    Input = [utils.path_join(drtrsfpath,f,dt='f') for f in filelist if f.endswith('.txt')]
    
    with open(Input[0], 'r') as f:
        line1 = f.readline()
    lit=line1.split('=')[1]
    print(lit)
    n=int(lit)
    
    n_split=3+n
    
    # Initialize the 3D arrays
    num_files = len(Input)
    Xcoeffs = np.zeros((n_split, n_split, num_files))  # Assuming each file has 5x5 Coeffs
    Ycoeffs = np.zeros((n_split, n_split, num_files))
    
    for k, filename in enumerate(Input):
        with open(filename, 'r') as f:
            lines = f.readlines()
    
        # Find the indices of the lines where the X and Y Coeffs start
        x_start = lines.index('X Coeffs -----------------------------------\n') + 2
        y_start = lines.index('Y Coeffs -----------------------------------\n') + 2
    
        # Read the X and Y Coeffs into ndarrays
        Xcoeffs[:,:,k] = np.array([list(map(float, line.split())) for line in lines[x_start-1:y_start-3]])
        Ycoeffs[:,:,k] = np.array([list(map(float, line.split())) for line in lines[y_start-1:y_start+4]])
    
    # Perform some operation on the Coeffs (multiplication by 2 in this example)
    for i in range(0,n_split):
        for j in range(0,n_split):
            # if not j == i == 0 or not j == i == n_split:
            Xcoeffs[i,j,:] =  gaussian_filter(utils.median_filter2(Xcoeffs[i,j,:],5),sigma=4,mode='nearest')
            Ycoeffs[j,i,:] = gaussian_filter(utils.median_filter2(Ycoeffs[j,i,:],5),sigma=4,mode='nearest')
    
    # Create output folder / clean previous execution.
    outputdir=utils.path_rectify(drtrsfpath.strip('/')+'_rectified',dt='d')
    os.makedirs(outputdir, exist_ok=True)
    prevexecution = [f for f in os.listdir(outputdir) if f.lower().endswith(('.txt'))]
    for f in prevexecution:
        os.remove(utils.path_join(outputdir)+f)
    Output = [utils.path_join(outputdir,f,dt='f') for f in filelist if f.endswith('.txt')]
    for k, filename in enumerate(Output):
        # Write the modified data back to the file
        with open(filename, 'w') as f:
            f.write(lines[0]+'\n')
            f.write('X Coeffs -----------------------------------\n')
            for row in Xcoeffs[:,:,k]:
                f.write(' '.join([f'{item:>20}' for item in map(str, row)]) + ' \n')
            f.write('\n')
            f.write('Y Coeffs -----------------------------------\n')
            for row in Ycoeffs[:,:,k]:
                f.write(' '.join([f'{item:>20}' for item in map(str, row)]) + ' \n')
                
    return outputdir

if __name__ == '__main__':    
    ## Image processing
    # test_IJ_bUnwarpJ()
    # test_image_threshold_hue()
    # test_rank_pixels()
    # test_images_outliers2nan()
    # test_image_adjust_contrast()
    # test_array_find_contrast_count_based()
    # test_utils.image_positivehist()
    # test_images_calculate_colocalisation()
    # test_images_vary_treshold()
    # test_images_scattergram()
    # test_PEEM_compare2i()
    # test_update_dataframe()
    # test_imlist_update_excel_sheet()
    # test_image_rankpixels()
    # test_PEEM_compare2i_2((40, 40), (60, 60))
    # test_PEEM_compare2i_2((45, 45), (55, 55))
    # test_PEEM_compare2i_2((47, 47), (53, 53))
    # test_PEEM_compare2i_3()
    # t0=time.time()
    # test_PEEM_compare2i_3(typeIm = 'peak')
    # t1=time.time()
    # print(f'Runtime test_compare2stacks3: {t1-t0}s')
    # test_PEEM_compare2i_4()

    print(os.getcwd())
